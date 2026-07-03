"""
Stage F — Critical-Path Delay Ledger  (Methodology B)
======================================================
Reconstructs, for every weekly snapshot, the driving path to the finish
milestone, identifies the controlling construction activity, and converts the
week-over-week movement of the forecast finish into a delay ledger attributed
to whatever was driving the finish (contemporaneous "windows" attribution).

Built entirely from the predecessor logic network — not from baseline
comparison. The number of snapshots is variable; nothing assumes a fixed count.

Per snapshot:
  1. Locate the finish milestone (config schedule.finish_milestone_task_name;
     fall back to the latest-finishing leaf if absent in early files).
  2. Trace the driving path backward: at each task follow the driving
     predecessor (the latest-finishing among those MS Project flags driving;
     if none is flagged, the latest-finishing predecessor). A visited-set
     prevents loops.
  3. Controlling activity = the earliest-finishing CONSTRUCTION task on the
     path that is not yet complete as of that snapshot. Record name, bucket,
     resource(s).

Delay ledger (between consecutive snapshots, in chronological order):
  - day change = calendar-day difference of the driving-path finish date.
  - attributed to the controlling activity at the LATER snapshot.
  - aggregated by resource (net days each sub controlled) and chronologically
    (the waterfall: consecutive windows under the same controlling activity
    grouped into one control-period).
  - totals: gross days added, days recovered, net.

Snapshot ordering precedence (most robust first):
  1. --manifest CSV  (stem,date) explicit override
  2. project status date  (captured by Stage C from the MPP data date)
  3. date parsed from the snapshot filename stem
  4. file modification time  (always available; last resort)
Same-date snapshots are de-duplicated, keeping one (the largest file).

Inputs:
  - project_config.json
  - Stage C output: output_root/stage_c/snapshots/<stem>.parquet
                    output_root/stage_c/predecessors/<stem>.parquet

Outputs (output_root/stage_f/):
  - snapshot_controlling.parquet   — per-snapshot controlling activity + finish
  - controlling_timeline.parquet   — snapshot runs grouped by controlling activity
  - delay_ledger_windows.parquet   — per consecutive-pair day change + attribution
  - waterfall_periods.parquet      — window runs grouped into control-periods
  - by_resource_net.parquet        — net days per controlling resource
  - delay_ledger_report.json       — totals (gross/recovered/net), counts
  - Delay_Ledger_<latest>.xlsx     — styled timeline + ledger + by-resource

No Java required — reads the Stage C parquet directly.

Usage:
  python critical_path.py [--config project_config.json] [--manifest order.csv]
"""

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------------
# Config + snapshot discovery / ordering
# ---------------------------------------------------------------------------

def load_config(config_path: str) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def parse_segments(outline_number) -> list[str]:
    """Split an MS Project outline number ('1.1.2.3') into segments."""
    if outline_number is None:
        return []
    s = str(outline_number).strip()
    if s == "" or s.lower() == "none" or s.lower() == "nan":
        return []
    return s.split(".")


def is_buyout_outline(outline_number, buyout_prefixes) -> bool:
    """True if outline_number equals, or is a descendant of, ANY prefix in
    buyout_prefixes. Segment-aware (so outline '1.23' never wrongly matches
    prefix '1.2' the way a naive string .startswith() would) and supports
    multiple disjoint buyout branches, since buyout packages are often
    added to the schedule after construction work already exists, so they
    don't all sit under one contiguous UID range or even one outline branch."""
    segs = parse_segments(outline_number)
    if not segs:
        return False
    for prefix in buyout_prefixes:
        pseg = parse_segments(prefix)
        if pseg and segs[:len(pseg)] == pseg:
            return True
    return False


_DATE_PATTERNS = [
    # (regex, (year_group, month_group, day_group), year_is_2digit)
    # Separators allow 1-2 chars: real snapshot files carry typos like
    # "3.19..26" (double dot), which a single-separator pattern misses —
    # and a missed filename date silently falls back to file-modification
    # time, which is meaningless after Stage B's COM export re-saves every
    # file (observed: 98 New Town files all re-saved on the same day, so
    # both LastSaved and mtime carry the export date, not the snapshot date).
    (re.compile(r"(20\d{2})[-._]{1,2}(\d{1,2})[-._]{1,2}(\d{1,2})"), (1, 2, 3), False),
    (re.compile(r"(\d{1,2})[-._]{1,2}(\d{1,2})[-._]{1,2}(20\d{2})"), (3, 1, 2), False),
    (re.compile(r"(\d{1,2})[-._]{1,2}(\d{1,2})[-._]{1,2}(\d{2})(?!\d)"), (3, 1, 2), True),
]


def parse_stem_date(stem: str):
    """Extract a date from a snapshot filename stem. Returns a date or None."""
    for rx, (yg, mg, dg), two_digit in _DATE_PATTERNS:
        m = rx.search(stem)
        if not m:
            continue
        try:
            y = int(m.group(yg))
            if two_digit:
                y += 2000
            mo = int(m.group(mg))
            d = int(m.group(dg))
            if 2000 <= y <= 2040 and 1 <= mo <= 12 and 1 <= d <= 31:
                return date(y, mo, d)
        except (ValueError, IndexError):
            continue
    return None


def select_single_snapshot(cfg: dict, snapshot_arg, parquets):
    """Choose ONE snapshot parquet for the single-snapshot stages (D/E/H).

    Precedence:
      1. explicit --snapshot CLI arg (exact stem match),
      2. schedule.analysis_snapshot in config (exact stem match; ignored if absent),
      3. the chronologically latest snapshot by parsed filename date,
      4. fallback: alphabetically last (only if no filename dates parse).

    `parquets` is a list of Path. Returns the chosen Path. Raises if a
    name is explicitly requested but not found.
    """
    by_stem = {p.stem: p for p in parquets}
    if snapshot_arg:
        if snapshot_arg in by_stem:
            return by_stem[snapshot_arg]
        raise FileNotFoundError(
            f"Snapshot '{snapshot_arg}' not found. Available: {sorted(by_stem)}")
    chosen = (cfg.get("schedule", {}) or {}).get("analysis_snapshot", "")
    if chosen and chosen in by_stem:
        return by_stem[chosen]
    dated = [(p, parse_stem_date(p.stem)) for p in parquets]
    if any(d is not None for _, d in dated):
        return max(dated, key=lambda t: (t[1] is not None, t[1] or date.min))[0]
    return sorted(parquets, key=lambda p: p.stem)[-1]


def discover_snapshots(cfg: dict, manifest_path: str | None):
    """
    Return an ordered, de-duplicated list of dicts:
        {stem, tasks_path, preds_path, date}
    """
    output_root = Path(cfg["paths"]["output_root"])
    snap_dir = output_root / "stage_c" / "snapshots"
    pred_dir = output_root / "stage_c" / "predecessors"
    if not snap_dir.exists():
        raise FileNotFoundError(
            f"Stage C output not found at {snap_dir}. Run extract_snapshots.py first."
        )

    parquets = sorted(snap_dir.glob("*.parquet"), key=lambda p: p.stem)
    if not parquets:
        raise FileNotFoundError(f"No snapshot parquet files in {snap_dir}.")

    # explicit manifest dates
    manifest_dates = {}
    if manifest_path:
        mdf = pd.read_csv(manifest_path)
        cols = {c.lower(): c for c in mdf.columns}
        scol = cols.get("stem") or cols.get("snapshot") or mdf.columns[0]
        dcol = cols.get("date") or mdf.columns[1]
        for _, r in mdf.iterrows():
            manifest_dates[str(r[scol])] = pd.Timestamp(r[dcol]).date()

    snaps = []
    used_mtime = False
    for p in parquets:
        stem = p.stem
        # ordering key precedence:  status date  >  filename date  >  modified time
        # `sort_ts` is the full-resolution key (date-based keys land at midnight,
        # so same-date snapshots de-duplicate; mtime keeps sub-day precision so
        # genuinely distinct same-day files are preserved).
        d = manifest_dates.get(stem)
        source = "manifest"
        if d is None:
            # 1) project status date (data date) captured by Stage C
            try:
                head = pd.read_parquet(p, columns=["status_date"])
                if not head.empty and pd.notna(head["status_date"].iloc[0]):
                    d = pd.Timestamp(head["status_date"].iloc[0]).date()
                    source = "status_date"
            except Exception:
                pass
        if d is None:
            # 2) date parsed from the filename stem
            d = parse_stem_date(stem)
            if d is not None:
                source = "filename"

        if d is not None:
            sort_ts = pd.Timestamp(d)              # midnight — same-date keys are equal
            disp_date = d
        else:
            # 3) file modification time (always available; sub-day precision retained)
            sort_ts = pd.Timestamp.fromtimestamp(p.stat().st_mtime)
            disp_date = sort_ts.date()
            source = "modified"
            used_mtime = True

        snaps.append({
            "stem": stem,
            "tasks_path": p,
            "preds_path": pred_dir / f"{stem}.parquet",
            "date": disp_date,
            "sort_ts": sort_ts,
            "date_source": source,
            "rows": p.stat().st_size,  # proxy for "largest file" dedup
        })

    if used_mtime:
        print("  NOTE: some snapshots had no status date or filename date — "
              "ordered those by file modification time. Provide --manifest for exact control.")

    # de-duplicate identical keys, keeping the largest file. Date-based keys at
    # midnight collapse same-date snapshots (per the brief); distinct mtimes don't.
    by_key = {}
    for s in snaps:
        key = s["sort_ts"]
        if key not in by_key or s["rows"] > by_key[key]["rows"]:
            by_key[key] = s
    deduped = sorted(by_key.values(), key=lambda s: s["sort_ts"])

    dropped = len(snaps) - len(deduped)
    if dropped:
        print(f"  De-duplicated {dropped} same-date snapshot(s), keeping the largest each.")

    return deduped


# ---------------------------------------------------------------------------
# Date coercion + finish helpers
# ---------------------------------------------------------------------------

def to_ts(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        ts = pd.Timestamp(v)
        return None if pd.isna(ts) else ts
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Driving-path trace  (pure, testable)
# ---------------------------------------------------------------------------

def trace_driving_path(start_uid, finish_of, preds_map, exists):
    """
    Walk backward from start_uid following the driving predecessor.

    finish_of(uid)  -> pd.Timestamp or None   (scheduled/forecast finish)
    preds_map[uid]  -> list of (pred_uid, is_driving)
    exists(uid)     -> bool (uid present in this snapshot's task table)

    Returns the ordered chain [start_uid, ..., earliest task].
    """
    path = [start_uid]
    visited = {start_uid}
    current = start_uid
    MIN = pd.Timestamp.min

    while True:
        preds = preds_map.get(current, [])
        # candidate pool: those MS Project flags driving, else all
        driving = [pu for (pu, drv) in preds if drv is True]
        pool = driving if driving else [pu for (pu, _) in preds]
        # restrict to tasks present this snapshot and not yet visited
        pool = [pu for pu in pool if exists(pu) and pu not in visited]
        if not pool:
            break
        nxt = max(pool, key=lambda u: (finish_of(u) or MIN))
        path.append(nxt)
        visited.add(nxt)
        current = nxt

    return path


def find_controlling(path, is_construction, is_complete, finish_of):
    """
    Controlling activity = earliest-finishing CONSTRUCTION task on the path
    that is not yet complete. Returns the uid, or None.

    is_construction(uid) -> bool   (construction-scope leaf, not summary)
    is_complete(uid)     -> bool   (pct_complete >= threshold)
    finish_of(uid)       -> pd.Timestamp or None
    """
    MAX = pd.Timestamp.max
    candidates = [u for u in path if is_construction(u) and not is_complete(u)]
    if not candidates:
        return None
    return min(candidates, key=lambda u: (finish_of(u) or MAX))


def find_concurrent(path, primary_uid, is_construction, is_complete, finish_of,
                    threshold_days=3):
    """
    Detect near-critical concurrency: incomplete construction tasks on the path
    whose finish lands within `threshold_days` of the primary controller's
    finish — i.e. running in parallel and nearly as constraining.

    This does NOT trace a second full path; it flags that the finish was being
    pressured from more than one direction at this snapshot, which the single
    driving-path trace would otherwise hide.

    Returns a list of concurrent uids (excluding the primary), ordered by finish.
    """
    if primary_uid is None:
        return []
    pf = finish_of(primary_uid)
    if pf is None:
        return []
    pf = pd.Timestamp(pf)
    hits = []
    for u in path:
        if u == primary_uid:
            continue
        if not is_construction(u) or is_complete(u):
            continue
        f = finish_of(u)
        if f is None:
            continue
        if abs((pd.Timestamp(f) - pf).days) <= threshold_days:
            hits.append((u, pd.Timestamp(f)))
    hits.sort(key=lambda t: t[1])
    return [u for u, _ in hits]


# ---------------------------------------------------------------------------
# Per-snapshot processing
# ---------------------------------------------------------------------------

def process_snapshot(snap, cfg, bucket_resolver):
    """Returns a dict describing the controlling state for one snapshot."""
    buyout_prefixes = cfg["schedule"]["buyout_outline_prefixes"]
    milestone_name = cfg["schedule"]["finish_milestone_task_name"].strip().lower()
    complete_thresh = cfg["schedule"].get("percent_complete_threshold_complete", 100)

    tasks = pd.read_parquet(snap["tasks_path"])
    try:
        preds = pd.read_parquet(snap["preds_path"])
    except (FileNotFoundError, OSError):
        preds = pd.DataFrame(columns=["task_uid", "pred_task_uid", "is_driving"])

    # ── task lookups ──────────────────────────────────────────────────────
    finish_map, pct_map, summary_map, name_map, res_map, parent_map, outline_map = \
        {}, {}, {}, {}, {}, {}, {}
    for row in tasks.itertuples(index=False):
        uid = int(row.uid)
        # forecast finish: scheduled, else actual, else baseline
        f = to_ts(getattr(row, "sched_finish", None)) \
            or to_ts(getattr(row, "actual_finish", None)) \
            or to_ts(getattr(row, "baseline_finish", None))
        finish_map[uid] = f
        pct_map[uid] = float(row.pct_complete) if pd.notna(getattr(row, "pct_complete", None)) else 0.0
        summary_map[uid] = bool(getattr(row, "is_summary", False))
        name_map[uid] = str(row.name) if pd.notna(row.name) else ""
        res_map[uid] = str(getattr(row, "resources", "") or "")
        p = getattr(row, "parent_uid", None)
        parent_map[uid] = int(p) if pd.notna(p) else None
        outline_map[uid] = getattr(row, "outline_number", None)

    # ── predecessor map ───────────────────────────────────────────────────
    preds_map = {}
    for row in preds.itertuples(index=False):
        tu = int(row.task_uid)
        pu = getattr(row, "pred_task_uid", None)
        if pd.isna(pu):
            continue
        pu = int(pu)
        drv = getattr(row, "is_driving", None)
        drv = bool(drv) if (drv is True or drv is False) else (
            True if str(drv).lower() == "true" else (False if str(drv).lower() == "false" else None))
        preds_map.setdefault(tu, []).append((pu, drv))

    # ── locate finish milestone ───────────────────────────────────────────
    milestone_uid = None
    cand = [u for u, nm in name_map.items() if nm.strip().lower() == milestone_name]
    if cand:
        milestone_uid = max(cand, key=lambda u: (finish_map.get(u) or pd.Timestamp.min))
    else:
        # fall back to the latest-finishing leaf
        leaves = [u for u in name_map if not summary_map.get(u)]
        if leaves:
            milestone_uid = max(leaves, key=lambda u: (finish_map.get(u) or pd.Timestamp.min))

    result = {
        "stem": snap["stem"],
        "date": snap["date"],
        "date_source": snap["date_source"],
        "milestone_uid": milestone_uid,
        "milestone_finish": None,
        "controlling_uid": None,
        "controlling_name": "",
        "controlling_bucket": "",
        "controlling_resources": "",
        "concurrent_count": 0,
        "concurrent_activities": "",
        "path_len": 0,
        "constr_on_path": 0,
    }
    if milestone_uid is None:
        return result

    result["milestone_finish"] = finish_map.get(milestone_uid)

    # ── trace driving path ────────────────────────────────────────────────
    path = trace_driving_path(
        milestone_uid,
        finish_of=lambda u: finish_map.get(u),
        preds_map=preds_map,
        exists=lambda u: u in finish_map,
    )
    result["path_len"] = len(path)

    def is_construction(u):
        # construction-scope work on the path, excluding the finish-milestone
        # anchor itself (it is the endpoint, not the next binding work)
        return (not is_buyout_outline(outline_map.get(u), buyout_prefixes)) \
            and (not summary_map.get(u, False)) and (u != milestone_uid)

    def is_complete(u):
        return pct_map.get(u, 0.0) >= complete_thresh

    result["constr_on_path"] = sum(1 for u in path if is_construction(u))

    controlling = find_controlling(
        path, is_construction, is_complete, lambda u: finish_map.get(u))

    if controlling is not None:
        result["controlling_uid"] = controlling
        result["controlling_name"] = name_map.get(controlling, "")
        res = res_map.get(controlling, "").strip()
        result["controlling_resources"] = res if res else "(unassigned)"
        result["controlling_bucket"] = bucket_resolver(controlling, name_map, parent_map)

        # ── concurrency detection ─────────────────────────────────────────
        threshold = cfg["critical_path"].get("concurrent_paths_threshold_days", 3) \
            if "critical_path" in cfg else 3
        concurrent = find_concurrent(
            path, controlling, is_construction, is_complete,
            lambda u: finish_map.get(u), threshold_days=threshold)
        result["concurrent_count"] = len(concurrent)
        # build a readable "Name (Resource)" list, deduped
        labels, seen = [], set()
        for u in concurrent:
            nm = name_map.get(u, "")
            rs = res_map.get(u, "").strip() or "(unassigned)"
            lbl = f"{nm} ({rs})"
            if lbl not in seen:
                seen.add(lbl)
                labels.append(lbl)
        result["concurrent_activities"] = "; ".join(labels)

    return result


# ---------------------------------------------------------------------------
# Construction bucket resolver (self-contained; mirrors Stage E)
# ---------------------------------------------------------------------------

def make_bucket_resolver(cfg):
    buckets = cfg["construction_variance"]["buckets"]
    bucket_lookup = {b.strip().lower(): b for b in buckets}

    def resolve(leaf_uid, name_map, parent_map):
        cur = leaf_uid
        seen = set()
        while cur is not None and cur not in seen:
            seen.add(cur)
            nm = name_map.get(cur, "").strip().lower()
            if nm in bucket_lookup:
                return bucket_lookup[nm]
            cur = parent_map.get(cur)
        return ""

    return resolve


# ---------------------------------------------------------------------------
# Ledger assembly  (pure, testable)
# ---------------------------------------------------------------------------

def build_ledger(snapshot_records):
    """
    From ordered per-snapshot records, build:
      windows_df, by_resource_df, waterfall_df, timeline_df, totals
    """
    recs = [r for r in snapshot_records if r["milestone_uid"] is not None]

    # ── windows (consecutive pairs) ───────────────────────────────────────
    windows = []
    for i in range(1, len(recs)):
        prev, curr = recs[i - 1], recs[i]
        pf, cf = prev["milestone_finish"], curr["milestone_finish"]
        delta = (pd.Timestamp(cf) - pd.Timestamp(pf)).days if (pf is not None and cf is not None) else 0
        windows.append({
            "from_date":            prev["date"],
            "to_date":              curr["date"],
            "from_finish":          pf,
            "to_finish":            cf,
            "day_change":           delta,
            "controlling_uid":      curr["controlling_uid"],
            "controlling_name":     curr["controlling_name"],
            "controlling_bucket":   curr["controlling_bucket"],
            "controlling_resources": curr["controlling_resources"],
            "concurrent_count":     curr.get("concurrent_count", 0),
            "concurrent_activities": curr.get("concurrent_activities", ""),
        })
    windows_df = pd.DataFrame(windows)

    # ── by resource ───────────────────────────────────────────────────────
    if not windows_df.empty:
        by_res = (windows_df.groupby("controlling_resources")["day_change"]
                  .sum().reset_index()
                  .rename(columns={"day_change": "net_days"})
                  .sort_values("net_days", key=lambda s: s.abs(), ascending=False)
                  .reset_index(drop=True))
    else:
        by_res = pd.DataFrame(columns=["controlling_resources", "net_days"])

    # ── waterfall: run-length over windows by controlling uid ─────────────
    wf_rows = []
    if not windows_df.empty:
        run_start = 0
        for i in range(1, len(windows) + 1):
            boundary = (i == len(windows)) or \
                       (windows[i]["controlling_uid"] != windows[run_start]["controlling_uid"]) or \
                       (windows[i]["controlling_uid"] is None and windows[run_start]["controlling_uid"] is None
                        and windows[i]["controlling_name"] != windows[run_start]["controlling_name"])
            if boundary:
                run = windows[run_start:i]
                net = sum(w["day_change"] for w in run)
                wf_rows.append({
                    "from_date":            run[0]["from_date"],
                    "to_date":              run[-1]["to_date"],
                    "controlling_name":     run[-1]["controlling_name"],
                    "controlling_bucket":   run[-1]["controlling_bucket"],
                    "controlling_resources": run[-1]["controlling_resources"],
                    "net_days":             net,
                    "end_finish":           run[-1]["to_finish"],
                })
                run_start = i
        # running total
        running = 0
        for r in wf_rows:
            running += r["net_days"]
            r["running_total"] = running
    waterfall_df = pd.DataFrame(wf_rows)

    # ── controlling timeline: run-length over SNAPSHOTS by controlling activity
    tl_rows = []
    if recs:
        run_start = 0
        for i in range(1, len(recs) + 1):
            boundary = (i == len(recs)) or \
                       (recs[i]["controlling_uid"] != recs[run_start]["controlling_uid"]) or \
                       (recs[i]["controlling_uid"] is None and recs[run_start]["controlling_uid"] is None
                        and recs[i]["controlling_name"] != recs[run_start]["controlling_name"])
            if boundary:
                run = recs[run_start:i]
                # any week in this run had concurrency?
                conc = next((r["concurrent_activities"] for r in reversed(run)
                             if r.get("concurrent_count", 0) > 0), "")
                tl_rows.append({
                    "from_date":            run[0]["date"],
                    "to_date":              run[-1]["date"],
                    "controlling_name":     run[-1]["controlling_name"],
                    "controlling_bucket":   run[-1]["controlling_bucket"],
                    "controlling_resources": run[-1]["controlling_resources"],
                    "driving_finish":       run[-1]["milestone_finish"],
                    "weeks":                len(run),
                    "concurrent_activities": conc,
                })
                run_start = i
    timeline_df = pd.DataFrame(tl_rows)

    # ── totals ────────────────────────────────────────────────────────────
    if not windows_df.empty:
        gross_added = int(windows_df.loc[windows_df["day_change"] > 0, "day_change"].sum())
        recovered = int(windows_df.loc[windows_df["day_change"] < 0, "day_change"].sum())
        net = int(windows_df["day_change"].sum())
    else:
        gross_added = recovered = net = 0

    concurrent_weeks = int(sum(1 for r in recs if r.get("concurrent_count", 0) > 0))

    totals = {
        "gross_days_added": gross_added,
        "days_recovered": recovered,
        "net_movement": net,
        "first_finish": str(recs[0]["milestone_finish"]) if recs else None,
        "last_finish": str(recs[-1]["milestone_finish"]) if recs else None,
        "snapshots_used": len(recs),
        "control_periods": len(wf_rows),
        "concurrent_weeks": concurrent_weeks,
    }

    return windows_df, by_res, waterfall_df, timeline_df, totals


# ---------------------------------------------------------------------------
# Styled workbook
# ---------------------------------------------------------------------------

NAVY = "FF1F4E78"; ZEBRA = "FFF2F6FB"; REDTINT = "FFFCE4E4"
GREENTINT = "FFE2EFDA"; GRAY = "FF606060"; WHITE = "FFFFFFFF"


def _fmt_date(v):
    ts = to_ts(v)
    return ts.strftime("%b %d") if ts is not None else ""


def write_workbook(path, cfg, totals, timeline_df, waterfall_df, by_res_df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    navy = PatternFill("solid", fgColor=NAVY)
    zebra = PatternFill("solid", fgColor=ZEBRA)
    red = PatternFill("solid", fgColor=REDTINT)
    green = PatternFill("solid", fgColor=GREENTINT)
    hf = Font(name="Calibri", bold=True, color=WHITE)
    tf = Font(name="Calibri", size=14, bold=True, color=NAVY)
    sf = Font(name="Calibri", size=10, color=GRAY)
    bf = Font(name="Calibri"); bb = Font(name="Calibri", bold=True)
    status = cfg["project"].get("analysis_status_date", "")

    def header(ws, title, sub, headers, hr=4):
        ws["A1"] = title; ws["A1"].font = tf
        ws["A2"] = sub; ws["A2"].font = sf
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=hr, column=c, value=h)
            cell.font = hf; cell.fill = navy
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def autosize(ws, headers, maxw=38):
        for c in range(1, len(headers) + 1):
            L = get_column_letter(c)
            mx = len(str(headers[c - 1]))
            for row in ws.iter_rows(min_col=c, max_col=c, min_row=5):
                for cell in row:
                    if cell.value is not None:
                        mx = max(mx, len(str(cell.value)))
            ws.column_dimensions[L].width = min(max(mx + 2, 9), maxw)

    wb = Workbook()

    # Timeline
    ws = wb.active; ws.title = "Controlling Timeline"
    hdrs = ["Period", "Controlling activity", "Resource", "Bucket", "Driving finish",
            "Weeks", "Concurrent (near-critical)"]
    header(ws, "Controlling Activity & Resource — Timeline",
           f"Driving-path control by snapshot. Status {status}.", hdrs)
    for i, (_, r) in enumerate(timeline_df.iterrows()):
        rr = 5 + i; fill = zebra if i % 2 else None
        period = _fmt_date(r["from_date"]) + ("–" + _fmt_date(r["to_date"])
                                              if r["from_date"] != r["to_date"] else "")
        vals = [period, r["controlling_name"], r["controlling_resources"],
                r["controlling_bucket"], _fmt_date(r["driving_finish"]), int(r["weeks"]),
                r.get("concurrent_activities", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = bf
            cell.alignment = Alignment(horizontal="center" if c in (1, 5, 6) else "left")
            if fill: cell.fill = fill
    autosize(ws, hdrs)

    # Waterfall / ledger
    ws = wb.create_sheet("Delay Ledger")
    hdrs = ["Period", "Controlling activity", "Resource", "Bucket",
            "Net days", "Running total", "Driving finish"]
    header(ws, "The Delay Ledger — How Many Days Each Handoff Added",
           f"Windows attribution. Net = gross added + recovered. Status {status}.", hdrs)
    for i, (_, r) in enumerate(waterfall_df.iterrows()):
        rr = 5 + i; fill = zebra if i % 2 else None
        period = _fmt_date(r["from_date"]) + "–" + _fmt_date(r["to_date"])
        vals = [period, r["controlling_name"], r["controlling_resources"],
                r["controlling_bucket"], int(r["net_days"]), int(r["running_total"]),
                _fmt_date(r["end_finish"])]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = bb if c == 5 else bf
            cell.alignment = Alignment(horizontal="center" if c in (1, 5, 6, 7) else "left")
            if c == 5:
                cell.fill = red if v > 0 else (green if v < 0 else (fill or PatternFill()))
            elif fill:
                cell.fill = fill
    # totals row
    tr = 5 + len(waterfall_df) + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = bb
    ws.cell(row=tr, column=5, value=f"+{totals['gross_days_added']} / {totals['days_recovered']}").font = bb
    ws.cell(row=tr, column=6, value=totals["net_movement"]).font = bb
    ws.cell(row=tr, column=6).fill = red if totals["net_movement"] > 0 else green
    autosize(ws, hdrs)

    # By resource
    ws = wb.create_sheet("By Resource")
    hdrs = ["Resource", "Net days controlling"]
    header(ws, "Net Calendar Days While Each Resource Controlled the Finish",
           f"Sum of day changes attributed to each controlling resource. Status {status}.", hdrs)
    for i, (_, r) in enumerate(by_res_df.iterrows()):
        rr = 5 + i; fill = zebra if i % 2 else None
        ws.cell(row=rr, column=1, value=r["controlling_resources"]).font = bf
        if fill: ws.cell(row=rr, column=1).fill = fill
        cell = ws.cell(row=rr, column=2, value=int(r["net_days"]))
        cell.font = bb
        cell.alignment = Alignment(horizontal="center")
        cell.fill = red if r["net_days"] > 0 else (green if r["net_days"] < 0 else (fill or PatternFill()))
    autosize(ws, hdrs)

    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Stage F — critical-path delay ledger")
    parser.add_argument("--config", default="project_config.json")
    parser.add_argument("--manifest", default=None,
                        help="Optional CSV (stem,date) to force snapshot ordering")
    args = parser.parse_args()

    cfg = load_config(args.config)
    project_name = cfg["project"]["name"]
    output_root = Path(cfg["paths"]["output_root"])
    stage_dir = output_root / "stage_f"
    stage_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"  Stage F — Critical-Path Delay Ledger")
    print(f"  Project : {project_name}")
    print(f"{'='*60}\n")

    snaps = discover_snapshots(cfg, args.manifest)
    print(f"  Snapshots (ordered, de-duplicated): {len(snaps)}")
    print(f"  Range: {snaps[0]['date']}  →  {snaps[-1]['date']}")
    print(f"  Ordering source: {snaps[0]['date_source']}\n")

    bucket_resolver = make_bucket_resolver(cfg)

    records = []
    for i, snap in enumerate(snaps, 1):
        rec = process_snapshot(snap, cfg, bucket_resolver)
        records.append(rec)
        fin = to_ts(rec["milestone_finish"])
        print(f"  [{i:>3}/{len(snaps)}] {snap['stem'][:34]:<34} "
              f"finish={fin.strftime('%Y-%m-%d') if fin else 'n/a':<10} "
              f"ctrl={rec['controlling_name'][:28]}")

    windows_df, by_res_df, waterfall_df, timeline_df, totals = build_ledger(records)

    # ── persist ───────────────────────────────────────────────────────────
    snap_df = pd.DataFrame([{k: v for k, v in r.items()} for r in records])
    snap_df.to_parquet(stage_dir / "snapshot_controlling.parquet", index=False)
    timeline_df.to_parquet(stage_dir / "controlling_timeline.parquet", index=False)
    windows_df.to_parquet(stage_dir / "delay_ledger_windows.parquet", index=False)
    waterfall_df.to_parquet(stage_dir / "waterfall_periods.parquet", index=False)
    by_res_df.to_parquet(stage_dir / "by_resource_net.parquet", index=False)

    xlsx_path = stage_dir / f"Delay_Ledger_{snaps[-1]['stem']}.xlsx"
    write_workbook(xlsx_path, cfg, totals, timeline_df, waterfall_df, by_res_df)

    report = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "project": project_name,
        "totals": totals,
        "snapshots_input": len(snaps),
        "ordering_source": snaps[0]["date_source"],
    }
    with open(stage_dir / "delay_ledger_report.json", "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)

    # ── console summary ───────────────────────────────────────────────────
    print(f"\n  {'─'*50}")
    print(f"  Forecast finish: {totals['first_finish']}  →  {totals['last_finish']}")
    print(f"  Gross added : +{totals['gross_days_added']} calendar days")
    print(f"  Recovered   : {totals['days_recovered']} calendar days")
    print(f"  Net movement: {'+' if totals['net_movement']>=0 else ''}{totals['net_movement']} calendar days")
    print(f"  Control-periods: {totals['control_periods']}")
    if totals.get("concurrent_weeks"):
        print(f"  Concurrent (near-critical) weeks: {totals['concurrent_weeks']} "
              f"— finish pressured from >1 direction (see timeline)")
    print(f"  (Harrison reference: +157 added, −70 recovered, +87 net)")

    print(f"\n{'='*60}")
    print(f"  Workbook : {xlsx_path}")
    print(f"  Parquet  : {stage_dir}")
    print(f"{'='*60}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
