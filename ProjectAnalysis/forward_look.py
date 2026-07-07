"""
Stage G — Forward Path-to-Completion Analysis  (Methodology C)
==============================================================
Reads the latest snapshot (plus the last N snapshots for the completion
range) and produces the four forward-looking views in Part IV of the brief.

1. Building turnover
   Each building's summary task: percent complete, baseline completion
   (saved baseline finish), forecast completion (current scheduled finish);
   slip = calendar-day difference. Feeds the lollipop chart in Stage K.

2. Float health
   Remaining (incomplete) construction tasks banded by MS Project total
   slack: Critical (<=0), then the bands defined in
   config.charting.float_health_bands_days. The clearest forward risk
   indicator (Harrison: 79% within 5 days of critical).

3. Completion range
   Span of the driving-path forecast finish over the last N weekly
   snapshots (config.charting.completion_range_lookback_weeks) — an
   empirical uncertainty band, not a modeled confidence interval. Prefers
   Stage F's forecast series; falls back to deriving it locally.

4. Look-ahead
   The not-yet-complete construction tasks on the current driving path —
   the forward mirror of Stage F's delay ledger. Intermediate markers
   (e.g. "BUILDING COMPLETE") are retained; only the Project Complete
   anchor is excluded.

Reuses Stage F (critical_path.py) for snapshot discovery/ordering, the
driving-path trace, and the construction bucket resolver, so the look-ahead
is literally "the same driving path used in Part II."

Inputs:
  - project_config.json
  - Stage C output (snapshots + predecessors)
  - Stage F output (optional, for the completion-range forecast series)

Outputs (output_root/stage_g/):
  - building_turnover.parquet
  - float_health.parquet
  - completion_range.json
  - look_ahead.parquet
  - forward_report.json
  - Forward_Look_<latest>.xlsx

Usage:
  python forward_look.py [--config project_config.json] [--manifest order.csv]
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

# Reuse Stage F utilities (same pipeline folder)
try:
    from critical_path import (discover_snapshots, trace_driving_path,
                               make_bucket_resolver, to_ts,
                               parse_segments, is_buyout_outline)
except ImportError as e:
    sys.exit("Stage G requires critical_path.py (Stage F) in the same folder. "
             f"Import failed: {e}")


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def load_config(config_path: str) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def status_from_pct(pct: float, complete_thresh: float) -> str:
    if pct >= complete_thresh:
        return "Complete"
    if pct <= 0:
        return "Not started"
    return "In progress"


# ---------------------------------------------------------------------------
# Shared lookups for one snapshot
# ---------------------------------------------------------------------------

def build_lookups(tasks_df: pd.DataFrame, preds_df: pd.DataFrame):
    finish_map, pct_map, summary_map, milestone_map = {}, {}, {}, {}
    name_map, res_map, parent_map, slack_map, outline_map = {}, {}, {}, {}, {}
    for row in tasks_df.itertuples(index=False):
        uid = int(row.uid)
        finish_map[uid] = (to_ts(getattr(row, "sched_finish", None))
                           or to_ts(getattr(row, "actual_finish", None))
                           or to_ts(getattr(row, "baseline_finish", None)))
        pct_map[uid] = float(row.pct_complete) if pd.notna(getattr(row, "pct_complete", None)) else 0.0
        summary_map[uid] = bool(getattr(row, "is_summary", False))
        milestone_map[uid] = bool(getattr(row, "is_milestone", False))
        name_map[uid] = str(row.name) if pd.notna(row.name) else ""
        res_map[uid] = str(getattr(row, "resources", "") or "")
        p = getattr(row, "parent_uid", None)
        parent_map[uid] = int(p) if pd.notna(p) else None
        s = getattr(row, "total_slack", None)
        slack_map[uid] = float(s) if pd.notna(s) else None
        outline_map[uid] = getattr(row, "outline_number", None)

    preds_map = {}
    for row in preds_df.itertuples(index=False):
        tu = int(row.task_uid)
        pu = getattr(row, "pred_task_uid", None)
        if pd.isna(pu):
            continue
        drv = getattr(row, "is_driving", None)
        drv = True if drv is True or str(drv).lower() == "true" else (
            False if drv is False or str(drv).lower() == "false" else None)
        preds_map.setdefault(tu, []).append((int(pu), drv))

    return dict(finish=finish_map, pct=pct_map, summary=summary_map,
                milestone=milestone_map, name=name_map, res=res_map,
                parent=parent_map, slack=slack_map, preds=preds_map,
                outline=outline_map)


def find_milestone_uid(lk, milestone_name: str):
    cand = [u for u, nm in lk["name"].items() if nm.strip().lower() == milestone_name]
    if cand:
        return max(cand, key=lambda u: (lk["finish"].get(u) or pd.Timestamp.min))
    leaves = [u for u in lk["name"] if not lk["summary"].get(u)]
    if leaves:
        return max(leaves, key=lambda u: (lk["finish"].get(u) or pd.Timestamp.min))
    return None


# ---------------------------------------------------------------------------
# 1. Building turnover
# ---------------------------------------------------------------------------

def build_turnover(tasks_df: pd.DataFrame, cfg: dict):
    complete_thresh = cfg["schedule"].get("percent_complete_threshold_complete", 100)
    building_names = cfg["buildings"]["names"]
    buyout_prefixes = cfg["schedule"]["buyout_outline_prefixes"]
    bld_phase = {}
    for ph in cfg["buildings"].get("phases", []):
        for b in ph.get("buildings", []):
            bld_phase[b] = ph.get("label", f"Phase {ph.get('phase_id','')}")

    # Building lookups must exclude the buyout branch: buyout packages group
    # their activities by building, so a schedule can carry DOZENS of summary
    # tasks named e.g. "Building 1" under buyout (New Town: 53 matches, 52 of
    # them buyout). Taking the first row by order silently returned a buyout
    # sub-group — 100% complete with a 2025 finish — instead of the one real
    # construction building summary.
    scope = tasks_df[~tasks_df["outline_number"].astype(str).map(
        lambda s: is_buyout_outline(s, buyout_prefixes))]

    rows, missing = [], []
    for bname in building_names:
        sub = scope[(scope["name"].astype(str).str.strip().str.lower()
                     == bname.strip().lower()) & (scope["is_summary"] == True)]
        if sub.empty:
            missing.append(bname)
            continue
        r = sub.iloc[0]
        pct = float(r["pct_complete"]) if pd.notna(r["pct_complete"]) else 0.0
        # Construction-scope building: prefer the construction baseline slot; the
        # generic baseline_finish is slot 0, empty when that slot is buyout-only.
        bf = to_ts(r.get("construction_baseline_finish")) or to_ts(r.get("baseline_finish"))
        ff = to_ts(r.get("sched_finish")) or to_ts(r.get("actual_finish"))
        slip = (pd.Timestamp(ff) - pd.Timestamp(bf)).days if (bf is not None and ff is not None) else None
        rows.append({
            "building":        bname,
            "phase":           bld_phase.get(bname, ""),
            "pct_complete":    pct,
            "baseline_finish": bf,
            "forecast_finish": ff,
            "slip_days":       slip,
            "status":          status_from_pct(pct, complete_thresh),
        })
    # Explicit columns so a project where NO configured building name matches
    # a schedule summary task (rows stays empty) still produces a correctly
    # shaped frame. pd.DataFrame([]) drops all columns, and this frame is
    # written to parquet unconditionally in main() — Stage K's
    # _data_building_lollipop then reads it back and does
    # df["baseline_finish"].notna(), which raised KeyError on the columnless
    # version.
    turnover_cols = ["building", "phase", "pct_complete", "baseline_finish",
                      "forecast_finish", "slip_days", "status"]
    df = pd.DataFrame(rows, columns=turnover_cols)
    return df, missing


# ---------------------------------------------------------------------------
# 2. Float health
# ---------------------------------------------------------------------------

def build_float_health(tasks_df: pd.DataFrame, cfg: dict):
    buyout_prefixes = cfg["schedule"]["buyout_outline_prefixes"]
    complete_thresh = cfg["schedule"].get("percent_complete_threshold_complete", 100)
    boundaries = cfg["charting"].get("float_health_bands_days", [0, 5, 10, 20])

    is_construction_mask = ~tasks_df["outline_number"].apply(
        lambda s: is_buyout_outline(s, buyout_prefixes))
    incomplete = tasks_df[is_construction_mask
                          & (tasks_df["is_summary"] == False)
                          & (tasks_df["pct_complete"].fillna(0) < complete_thresh)].copy()

    # band definitions from the boundary list
    bands = [(f"≤{boundaries[0]} (Critical)", lambda s, b=boundaries[0]: s <= b)]
    for i in range(len(boundaries) - 1):
        lo, hi = boundaries[i], boundaries[i + 1]
        bands.append((f"{lo + 1}–{hi}", lambda s, lo=lo, hi=hi: lo < s <= hi))
    bands.append((f">{boundaries[-1]}", lambda s, b=boundaries[-1]: s > b))

    slacks = incomplete["total_slack"]
    valid = slacks.notna()
    unknown = int((~valid).sum())
    total = int(valid.sum())

    counts = []
    for label, pred in bands:
        c = int(slacks[valid].map(pred).sum())
        counts.append({"band": label, "count": c,
                       "pct_of_incomplete": round(c / total * 100, 1) if total else 0.0})
    band_df = pd.DataFrame(counts)

    taut_boundary = boundaries[1] if len(boundaries) > 1 else boundaries[0]
    taut_count = int((slacks[valid] <= taut_boundary).sum())
    taut_pct = round(taut_count / total * 100, 1) if total else 0.0

    summary = {
        "total_incomplete_with_slack": total,
        "unknown_slack_count": unknown,
        "within_{}_days_pct".format(taut_boundary): taut_pct,
        "taut_boundary_days": taut_boundary,
    }
    return band_df, summary


# ---------------------------------------------------------------------------
# 3. Completion range
# ---------------------------------------------------------------------------

def build_completion_range(cfg: dict, ordered_snaps: list, milestone_name: str):
    n = cfg["charting"].get("completion_range_lookback_weeks", 8)
    output_root = Path(cfg["paths"]["output_root"])
    stage_f_series = output_root / "stage_f" / "snapshot_controlling.parquet"

    series = []  # list of (date, forecast_finish)

    if stage_f_series.exists():
        sf = pd.read_parquet(stage_f_series)
        sf = sf[sf["milestone_finish"].notna()].copy()
        sf["date"] = pd.to_datetime(sf["date"])
        sf = sf.sort_values("date").tail(n)
        for _, r in sf.iterrows():
            series.append((r["date"].date(), to_ts(r["milestone_finish"])))
        source = "stage_f"
    else:
        # fallback: derive milestone finish from the last N snapshots directly
        for snap in ordered_snaps[-n:]:
            td = pd.read_parquet(snap["tasks_path"])
            try:
                pdf = pd.read_parquet(snap["preds_path"])
            except (FileNotFoundError, OSError):
                pdf = pd.DataFrame(columns=["task_uid", "pred_task_uid", "is_driving"])
            lk = build_lookups(td, pdf)
            muid = find_milestone_uid(lk, milestone_name)
            fin = lk["finish"].get(muid) if muid is not None else None
            series.append((snap["date"], fin))
        source = "local"

    finishes = [f for (_, f) in series if f is not None]
    if not finishes:
        return {"available": False, "source": source}

    earliest = min(finishes)
    latest = max(finishes)
    current = series[-1][1]

    # max week-to-week swing
    swing = 0
    for i in range(1, len(finishes)):
        swing = max(swing, abs((pd.Timestamp(finishes[i]) - pd.Timestamp(finishes[i - 1])).days))

    return {
        "available": True,
        "source": source,
        "snapshots_used": len(finishes),
        "lookback_weeks": n,
        "earliest_forecast": str(pd.Timestamp(earliest).date()),
        "latest_forecast": str(pd.Timestamp(latest).date()),
        "current_forecast": str(pd.Timestamp(current).date()) if current is not None else None,
        "range_days": (pd.Timestamp(latest) - pd.Timestamp(earliest)).days,
        "max_week_swing_days": swing,
        "series": [(str(d), str(pd.Timestamp(f).date()) if f is not None else None)
                   for (d, f) in series],
    }


# ---------------------------------------------------------------------------
# 4. Look-ahead
# ---------------------------------------------------------------------------

def build_look_ahead(tasks_df, preds_df, cfg, bucket_resolver):
    buyout_prefixes = cfg["schedule"]["buyout_outline_prefixes"]
    complete_thresh = cfg["schedule"].get("percent_complete_threshold_complete", 100)
    milestone_name = cfg["schedule"]["finish_milestone_task_name"].strip().lower()

    lk = build_lookups(tasks_df, preds_df)
    muid = find_milestone_uid(lk, milestone_name)
    if muid is None:
        return pd.DataFrame(), None

    path, _terminus = trace_driving_path(
        muid,
        finish_of=lambda u: lk["finish"].get(u),
        preds_map=lk["preds"],
        exists=lambda u: u in lk["finish"],
    )

    rows = []
    for u in path:
        if u == muid:
            continue
        if is_buyout_outline(lk["outline"].get(u), buyout_prefixes) or lk["summary"].get(u):
            continue
        if lk["pct"].get(u, 0.0) >= complete_thresh:
            continue
        rows.append({
            "forecast_finish": lk["finish"].get(u),
            "activity":        lk["name"].get(u, ""),
            "resources":       lk["res"].get(u, "").strip() or "(unassigned)",
            "bucket":          bucket_resolver(u, lk["name"], lk["parent"]),
            "total_slack":     lk["slack"].get(u),
            "pct_complete":    lk["pct"].get(u, 0.0),
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("forecast_finish", na_position="last").reset_index(drop=True)
    return df, muid


# ---------------------------------------------------------------------------
# 5. Per-building driving paths  (5.7 — superintendent/PM-facing)
# ---------------------------------------------------------------------------

def build_building_paths(tasks_df, preds_df, cfg, bucket_resolver):
    """Backward driving-path trace against each configured building's own
    finish (latest snapshot only): per-building controlling activity + a short
    look-ahead. This is what makes the per-building section actionable — the
    project-level path only names whichever building drives the overall
    finish; a superintendent needs *their* building's path."""
    building_names = cfg["buildings"]["names"]
    buyout_prefixes = cfg["schedule"]["buyout_outline_prefixes"]
    complete_thresh = cfg["schedule"].get("percent_complete_threshold_complete", 100)

    lk = build_lookups(tasks_df, preds_df)

    # children index for subtree walks
    children = {}
    for u, p in lk["parent"].items():
        if p is not None:
            children.setdefault(p, []).append(u)

    def subtree_leaves(root):
        out, stack = [], [root]
        while stack:
            cur = stack.pop()
            kids = children.get(cur, [])
            if not kids and cur != root and not lk["summary"].get(cur):
                out.append(cur)
            stack.extend(kids)
        return out

    rows = []
    for bname in building_names:
        # Same buyout-branch exclusion as build_turnover: buyout packages
        # contain summary tasks named after buildings; only the construction
        # branch's building summary is the real one.
        summ = [u for u, nm in lk["name"].items()
                if nm.strip().lower() == bname.strip().lower()
                and lk["summary"].get(u)
                and not is_buyout_outline(lk["outline"].get(u), buyout_prefixes)]
        if not summ:
            continue
        root = summ[0]
        leaves = subtree_leaves(root)
        if not leaves:
            continue
        # anchor = the building's own finish: its latest-finishing leaf
        # (a turnover milestone when the schedule has one, since that is by
        # definition the last thing in the building's subtree)
        anchor = max(leaves, key=lambda u: (lk["finish"].get(u) or pd.Timestamp.min))
        path, terminus = trace_driving_path(
            anchor,
            finish_of=lambda u: lk["finish"].get(u),
            preds_map=lk["preds"],
            exists=lambda u: u in lk["finish"],
        )
        incomplete = [u for u in path
                      if not lk["summary"].get(u)
                      and not is_buyout_outline(lk["outline"].get(u), buyout_prefixes)
                      and lk["pct"].get(u, 0.0) < complete_thresh]
        incomplete.sort(key=lambda u: (lk["finish"].get(u) or pd.Timestamp.max))
        ctl = incomplete[0] if incomplete else None
        look = " → ".join(lk["name"].get(u, "") for u in incomplete[:4])
        rows.append({
            "building":              bname,
            "anchor_activity":       lk["name"].get(anchor, ""),
            "anchor_finish":         lk["finish"].get(anchor),
            "controlling_activity":  lk["name"].get(ctl, "") if ctl else "(all complete)",
            "controlling_resources": (lk["res"].get(ctl, "").strip() or "(unassigned)") if ctl else "",
            "controlling_finish":    lk["finish"].get(ctl) if ctl else None,
            "remaining_on_path":     len(incomplete),
            "look_ahead":            look,
            "terminus_reason":       terminus,
        })
    cols = ["building", "anchor_activity", "anchor_finish", "controlling_activity",
            "controlling_resources", "controlling_finish", "remaining_on_path",
            "look_ahead", "terminus_reason"]
    return pd.DataFrame(rows, columns=cols)


# ---------------------------------------------------------------------------
# 6. S-curve + slip velocity  (5.8 — executive-facing)
# ---------------------------------------------------------------------------

def build_progress_curve(cfg, ordered_snaps):
    """Planned vs actual percent-complete over time (construction scope).

    ACTUAL %(snapshot) = that snapshot's own baseline-duration-weighted
    pct_complete — the progress the schedule itself reports (the dashboard
    'construction complete' headline reads the latest actual). This is NOT
    re-weighted to the baseline of record, so pinning a baseline doesn't move
    the reported % done.

    PLANNED %(d) integrates the baseline-of-RECORD window (the pinned/original
    baseline, per schedule.baseline_basis) up to d, so the planned line runs to
    100% at the baseline finish — independent of the snapshot's own saved
    baseline (which may be rebaselined or, for a buyout-only slot-0 baseline,
    carry no construction dates at all)."""
    buyout_prefixes = cfg["schedule"]["buyout_outline_prefixes"]
    output_root = cfg["paths"]["output_root"]
    cols = ["uid", "outline_number", "is_summary", "pct_complete",
            "baseline_start", "baseline_finish", "baseline_duration",
            "construction_baseline_start", "construction_baseline_finish",
            "construction_baseline_duration"]

    def coalesce_constr(d):
        # Construction scope uses the construction baseline slot; the generic
        # baseline_* columns are always slot 0. Prefer construction_* per leaf.
        for base_c, con_c in (("baseline_start",    "construction_baseline_start"),
                              ("baseline_finish",   "construction_baseline_finish"),
                              ("baseline_duration", "construction_baseline_duration")):
            if con_c in d.columns:
                d[base_c] = d[con_c].where(d[con_c].notna(), d[base_c])
        return d

    def constr_leaves(d):
        m = (~d["is_summary"]) & (~d["outline_number"].astype(str).map(
            lambda s: is_buyout_outline(s, buyout_prefixes)))
        return d[m]

    records = []
    last_snap = None
    for snap in ordered_snaps:
        try:
            df = pd.read_parquet(snap["tasks_path"], columns=cols)
        except Exception:
            continue
        df = coalesce_constr(df)                       # snapshot's OWN baseline
        wb = constr_leaves(df)
        wb = wb[wb["baseline_duration"].notna() & (wb["baseline_duration"] > 0)]
        if wb.empty:
            continue
        total = wb["baseline_duration"].sum()
        actual_pct = float((wb["pct_complete"].fillna(0) * wb["baseline_duration"]).sum() / total)
        records.append({"date": pd.Timestamp(snap["date"]), "actual_pct": actual_pct})
        last_snap = snap
    if not records or last_snap is None:
        return pd.DataFrame(columns=["date", "planned_pct", "actual_pct"])

    # PLANNED baseline = baseline of record applied to the latest snapshot's
    # construction leaves. apply_original_baselines is a no-op for basis
    # 'current', so the planned line then uses the snapshot's own baseline.
    from construction_variance import apply_original_baselines
    pdf = pd.read_parquet(last_snap["tasks_path"], columns=cols)
    pdf = apply_original_baselines(pdf, cfg, output_root, "construction_")
    pdf = apply_original_baselines(pdf, cfg, output_root, "")
    pdf = coalesce_constr(pdf)
    pb = constr_leaves(pdf)
    pb = pb[pb["baseline_duration"].notna() & (pb["baseline_duration"] > 0)]
    bs = pd.to_datetime(pb["baseline_start"])
    bf = pd.to_datetime(pb["baseline_finish"])
    dur = pb["baseline_duration"].astype(float)
    total_p = dur.sum()
    span_days = (bf - bs).dt.days.clip(lower=1)

    curve = pd.DataFrame(records)
    planned = []
    for d in curve["date"]:
        frac = ((d - bs).dt.days / span_days).clip(0, 1)
        planned.append(float((frac * dur).sum() / total_p * 100.0))
    curve["planned_pct"] = planned
    return curve[["date", "planned_pct", "actual_pct"]]


def build_slip_velocity(comp_range):
    """Trailing regression of forecast-finish movement over the completion-
    range window: 'at the current trend, completion projects to X.' This is
    an EMPIRICAL TREND read (same epistemics as the completion range), not a
    forecast — the framing matters and is repeated wherever it renders."""
    series = comp_range.get("series") or []
    pts = [(pd.Timestamp(d), pd.Timestamp(f)) for d, f in series if f is not None]
    if len(pts) < 3:
        return {"available": False}
    x = pd.Series([p[0].toordinal() for p in pts], dtype=float)
    y = pd.Series([p[1].toordinal() for p in pts], dtype=float)
    n = len(x)
    sx, sy = x.sum(), y.sum()
    sxx = (x * x).sum(); sxy = (x * y).sum()
    denom = n * sxx - sx * sx
    if denom == 0:
        return {"available": False}
    slope = (n * sxy - sx * sy) / denom          # finish days moved per calendar day
    intercept = (sy - slope * sx) / n
    out = {
        "available": True,
        "window_snapshots": n,
        "slope_days_per_week": round(slope * 7, 2),
    }
    if slope < 1.0:
        # fixed point where the trending forecast equals the calendar
        x_star = intercept / (1.0 - slope)
        try:
            out["projected_completion"] = str(pd.Timestamp.fromordinal(int(round(x_star))).date())
            out["diverging"] = False
        except (ValueError, OverflowError):
            out["diverging"] = True
    else:
        # forecast slipping >= 1 day per day — trend never converges
        out["diverging"] = True
    return out


# ---------------------------------------------------------------------------
# Styled workbook
# ---------------------------------------------------------------------------

NAVY = "FF1F4E78"; ZEBRA = "FFF2F6FB"; REDTINT = "FFFCE4E4"
GREENTINT = "FFE2EFDA"; GRAY = "FF606060"; WHITE = "FFFFFFFF"; AMBER = "FFFFF2CC"


def _d(v):
    ts = to_ts(v)
    return ts.strftime("%b %d") if ts is not None else ""


def write_workbook(path, cfg, turnover_df, band_df, float_summary, comp_range, look_df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    navy = PatternFill("solid", fgColor=NAVY); zebra = PatternFill("solid", fgColor=ZEBRA)
    red = PatternFill("solid", fgColor=REDTINT); green = PatternFill("solid", fgColor=GREENTINT)
    amber = PatternFill("solid", fgColor=AMBER)
    hf = Font(name="Calibri", bold=True, color=WHITE)
    tf = Font(name="Calibri", size=14, bold=True, color=NAVY)
    sf = Font(name="Calibri", size=10, color=GRAY)
    bf = Font(name="Calibri"); bb = Font(name="Calibri", bold=True)
    status = cfg["project"].get("analysis_status_date", "")

    def header(ws, title, sub, headers, hr=4):
        ws["A1"] = title; ws["A1"].font = tf
        ws["A2"] = sub; ws["A2"].font = sf
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=hr, column=c, value=h)
            cell.font = hf; cell.fill = navy
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def autosize(ws, headers, maxw=38):
        for c in range(1, len(headers) + 1):
            L = get_column_letter(c); mx = len(str(headers[c - 1]))
            for row in ws.iter_rows(min_col=c, max_col=c, min_row=5):
                for cell in row:
                    if cell.value is not None:
                        mx = max(mx, len(str(cell.value)))
            ws.column_dimensions[L].width = min(max(mx + 2, 9), maxw)

    wb = Workbook()

    # ── Building turnover ─────────────────────────────────────────────────
    ws = wb.active; ws.title = "Building Turnover"
    hdrs = ["Building", "Phase", "% Complete", "Baseline", "Forecast", "Slip (days)", "Status"]
    header(ws, "Building Turnover — Baseline vs Forecast",
           f"Per-building summary task. Slip = forecast − baseline (calendar days). Status {status}.", hdrs)
    tv = turnover_df.sort_values("slip_days", ascending=False, na_position="last") \
        if not turnover_df.empty else turnover_df
    for i, (_, r) in enumerate(tv.iterrows()):
        rr = 5 + i; fill = zebra if i % 2 else None
        vals = [r["building"], r["phase"],
                (r["pct_complete"] / 100.0), _d(r["baseline_finish"]),
                _d(r["forecast_finish"]),
                int(r["slip_days"]) if pd.notna(r["slip_days"]) else "", r["status"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = bf
            if c == 3:
                cell.number_format = "0%"; cell.alignment = Alignment(horizontal="center")
            elif c == 6:
                cell.font = bb; cell.alignment = Alignment(horizontal="center")
                if isinstance(v, int):
                    cell.fill = red if v > 0 else (green if v < 0 else (fill or PatternFill()))
            else:
                cell.alignment = Alignment(horizontal="center" if c in (4, 5, 7) else "left")
            if fill and not (c == 6 and isinstance(v, int)):
                cell.fill = fill
    autosize(ws, hdrs)

    # ── Float health ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Float Health")
    hdrs = ["Slack Band", "Task Count", "% of Incomplete"]
    tb = float_summary.get("taut_boundary_days")
    header(ws, "Float Health — Remaining Cushion",
           f"Incomplete construction tasks by total slack. "
           f"{float_summary.get(f'within_{tb}_days_pct')}% within {tb} days of critical. Status {status}.", hdrs)
    for i, (_, r) in enumerate(band_df.iterrows()):
        rr = 5 + i
        is_crit = "Critical" in str(r["band"])
        fill = amber if is_crit else (zebra if i % 2 else None)
        vals = [r["band"], int(r["count"]), r["pct_of_incomplete"] / 100.0]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = bb if is_crit else bf
            if c == 3:
                cell.number_format = "0.0%"; cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="center" if c == 2 else "left")
            if fill:
                cell.fill = fill
    autosize(ws, hdrs)

    # ── Completion range ──────────────────────────────────────────────────
    ws = wb.create_sheet("Completion Range")
    ws["A1"] = "Completion Forecast — A Range, Not a Point"; ws["A1"].font = tf
    ws["A2"] = f"Driving-path forecast finish over the last {comp_range.get('lookback_weeks','')} snapshots. Status {status}."
    ws["A2"].font = sf
    if comp_range.get("available"):
        pairs = [
            ("Current forecast", comp_range["current_forecast"]),
            ("Earliest in window", comp_range["earliest_forecast"]),
            ("Latest in window", comp_range["latest_forecast"]),
            ("Range (days)", comp_range["range_days"]),
            ("Max week-to-week swing (days)", comp_range["max_week_swing_days"]),
            ("Snapshots used", comp_range["snapshots_used"]),
        ]
        for i, (k, v) in enumerate(pairs):
            rr = 4 + i
            kc = ws.cell(row=rr, column=1, value=k); kc.font = bb
            vc = ws.cell(row=rr, column=2, value=v); vc.font = bf
            vc.alignment = Alignment(horizontal="left")
        ws.column_dimensions["A"].width = 32; ws.column_dimensions["B"].width = 16
    else:
        ws["A4"] = "Completion range unavailable (no forecast series found)."

    # ── Look-ahead ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Look-Ahead")
    hdrs = ["#", "Forecast Finish", "Remaining controlling activity", "Resource",
            "Building", "Total Slack (d)", "% Complete"]
    header(ws, "Look-Ahead — What Controls the Remaining Work",
           f"Not-yet-complete construction tasks on the current driving path. Status {status}.", hdrs)
    for i, (_, r) in enumerate(look_df.iterrows()):
        rr = 5 + i; fill = zebra if i % 2 else None
        vals = [i + 1, _d(r["forecast_finish"]), r["activity"], r["resources"],
                r["bucket"],
                round(float(r["total_slack"]), 1) if pd.notna(r["total_slack"]) else "",
                (float(r["pct_complete"]) / 100.0)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = bf
            if c == 7:
                cell.number_format = "0%"; cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="center" if c in (1, 2, 6) else "left")
            if fill:
                cell.fill = fill
    autosize(ws, hdrs)

    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Stage G — forward path-to-completion")
    parser.add_argument("--config", default="project_config.json")
    parser.add_argument("--manifest", default=None)
    args = parser.parse_args()

    cfg = load_config(args.config)
    project_name = cfg["project"]["name"]
    output_root = Path(cfg["paths"]["output_root"])
    stage_dir = output_root / "stage_g"
    stage_dir.mkdir(parents=True, exist_ok=True)
    milestone_name = cfg["schedule"]["finish_milestone_task_name"].strip().lower()

    print(f"\n{'='*60}")
    print(f"  Stage G — Forward Path-to-Completion")
    print(f"  Project : {project_name}")
    print(f"{'='*60}\n")

    snaps = discover_snapshots(cfg, args.manifest)
    latest = snaps[-1]
    print(f"  Latest snapshot: {latest['stem']}  ({latest['date']})")
    print(f"  Snapshots available: {len(snaps)}\n")

    latest_tasks = pd.read_parquet(latest["tasks_path"])
    # baseline_basis='original': turnover baselines = each building's first-
    # ever-saved baseline finish, not the latest rebaselined value
    from construction_variance import apply_original_baselines
    latest_tasks = apply_original_baselines(latest_tasks, cfg, output_root, "")
    # Also apply the construction-scoped original baseline: building summaries
    # are construction scope, and when the generic (slot-0) baseline is
    # buyout-only their baseline_finish is empty. build_turnover falls back to
    # construction_baseline_finish, so populate it with the original plan too.
    latest_tasks = apply_original_baselines(latest_tasks, cfg, output_root, "construction_")
    try:
        latest_preds = pd.read_parquet(latest["preds_path"])
    except (FileNotFoundError, OSError):
        latest_preds = pd.DataFrame(columns=["task_uid", "pred_task_uid", "is_driving"])

    bucket_resolver = make_bucket_resolver(cfg)

    # 1. turnover
    turnover_df, missing = build_turnover(latest_tasks, cfg)
    # 2. float health
    band_df, float_summary = build_float_health(latest_tasks, cfg)
    # 3. completion range
    comp_range = build_completion_range(cfg, snaps, milestone_name)
    # 4. look-ahead
    look_df, look_muid = build_look_ahead(latest_tasks, latest_preds, cfg, bucket_resolver)
    # 5. per-building driving paths (5.7)
    bpath_df = build_building_paths(latest_tasks, latest_preds, cfg, bucket_resolver)
    # 6. S-curve + slip velocity (5.8)
    curve_df = build_progress_curve(cfg, snaps)
    velocity = build_slip_velocity(comp_range)

    # ── persist ───────────────────────────────────────────────────────────
    turnover_df.to_parquet(stage_dir / "building_turnover.parquet", index=False)
    band_df.to_parquet(stage_dir / "float_health.parquet", index=False)
    if not look_df.empty:
        look_df.to_parquet(stage_dir / "look_ahead.parquet", index=False)
    bpath_df.to_parquet(stage_dir / "building_driving_paths.parquet", index=False)
    curve_df.to_parquet(stage_dir / "progress_curve.parquet", index=False)
    with open(stage_dir / "completion_range.json", "w", encoding="utf-8") as f:
        json.dump(comp_range, f, indent=2, default=str)

    xlsx_path = stage_dir / f"Forward_Look_{latest['stem']}.xlsx"
    write_workbook(xlsx_path, cfg, turnover_df, band_df, float_summary, comp_range, look_df)

    report = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "project": project_name,
        "latest_snapshot": latest["stem"],
        "buildings_reported": len(turnover_df),
        "buildings_missing": missing,
        "float_summary": float_summary,
        "completion_range": {k: v for k, v in comp_range.items() if k != "series"},
        "look_ahead_count": len(look_df),
        "building_paths_count": len(bpath_df),
        "slip_velocity": velocity,
        "progress_curve_points": len(curve_df),
    }
    with open(stage_dir / "forward_report.json", "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)

    # ── console summary ───────────────────────────────────────────────────
    print("  Building turnover:")
    if not turnover_df.empty:
        sl = turnover_df["slip_days"].dropna()
        if len(sl):
            print(f"    {len(turnover_df)} buildings | slip range +{int(sl.min())} to +{int(sl.max())} days")
    if missing:
        print(f"    WARNING: building summary not found for: {missing}")

    tb = float_summary.get("taut_boundary_days")
    print(f"\n  Float health: {float_summary[f'within_{tb}_days_pct']}% within {tb} days of critical "
          f"({float_summary['total_incomplete_with_slack']} incomplete tasks)")
    print(f"    (Harrison reference: 79% within 5 days)")
    for _, r in band_df.iterrows():
        print(f"    {r['band']:<16} {int(r['count']):>5}  ({r['pct_of_incomplete']}%)")

    if comp_range.get("available"):
        print(f"\n  Completion range (last {comp_range['snapshots_used']} snapshots): "
              f"{comp_range['earliest_forecast']} → {comp_range['latest_forecast']} "
              f"(span {comp_range['range_days']}d, swing {comp_range['max_week_swing_days']}d)")
        print(f"    Current forecast: {comp_range['current_forecast']}")

    print(f"\n  Look-ahead: {len(look_df)} remaining tasks on the driving path")
    if not look_df.empty:
        for _, r in look_df.head(8).iterrows():
            print(f"    {_d(r['forecast_finish']):<8} {r['activity'][:34]:<34} "
                  f"{r['resources'][:16]:<16} {r['bucket']}")

    if not bpath_df.empty:
        print(f"\n  Per-building driving paths ({len(bpath_df)}):")
        for _, r in bpath_df.iterrows():
            print(f"    {r['building'][:22]:<24} controls: {r['controlling_activity'][:32]:<34} "
                  f"{r['controlling_resources'][:16]}")

    if velocity.get("available"):
        if velocity.get("diverging"):
            print(f"\n  Slip velocity: {velocity['slope_days_per_week']:+.1f} d/wk over last "
                  f"{velocity['window_snapshots']} snapshots — trend does not converge (slipping ≥1 day/day)")
        else:
            print(f"\n  Slip velocity: {velocity['slope_days_per_week']:+.1f} d/wk — at current trend, "
                  f"completion projects to {velocity['projected_completion']} (empirical trend, not a forecast)")

    print(f"\n{'='*60}")
    print(f"  Workbook : {xlsx_path}")
    print(f"  Parquet  : {stage_dir}")
    print(f"{'='*60}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
