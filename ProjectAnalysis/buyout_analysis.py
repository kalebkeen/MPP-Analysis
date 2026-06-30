"""
Stage H — Buyout Phase Analysis  (Methodology D)
================================================
Self-contained analysis of the procurement and subcontracting workflow.
Reads the Stage D structural grouping and the Stage C snapshot, classifies
every buyout leaf into a workflow stage, and measures variance on two bases.

The critical methodological move (the summary-span vs leaf-sum trap):
  A package's headline baseline/actual come from the package SUMMARY TASK's own
  MS Project duration — its scheduled working span. Leaf durations are NEVER
  summed: activities like "Order Drafted" recur once per building (~16 copies),
  so summing would inflate a package's total by the building count.

Two measurements (the two-measurement framework):
  - Duration variance = Actual − Baseline duration (working days).
  - Start-date slip    = Actual start − Baseline start (Mon-Fri business days).
  Both are computed per leaf; the workbook reports duration, and start-slip is
  persisted for the narrative (Stage J).

Stage classification is config-driven and priority-ordered
(config.buyout_analysis.stage_classification): the first stage whose keyword
(case-insensitive substring) appears in the activity name wins. This reproduces
the Harrison workbook's six-category taxonomy exactly.

Inputs:
  - project_config.json
  - Stage D output: output_root/stage_d/buyout_grouping.parquet
                    output_root/stage_d/buyout_packages.parquet
  - Stage C output: output_root/stage_c/snapshots/<snapshot>.parquet

Outputs (output_root/stage_h/):
  - buyout_summary.parquet         — Section × Group bucket totals
  - buyout_packages_ranked.parquet — package-level (summary-span) metrics
  - buyout_stage_breakdown.parquet — workflow-stage roll-up
  - buyout_activity_detail.parquet — every leaf, with stage + start-slip
  - buyout_report.json
  - Buyout_Duration_Variance_<snapshot>.xlsx  — J New Town styled workbook

No Java required — reads the Stage C / Stage D parquet directly.

Usage:
  python buyout_analysis.py [--config project_config.json] [--snapshot <stem>]
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------------
# Config + snapshot selection
# ---------------------------------------------------------------------------

def load_config(config_path: str) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def find_snapshot(cfg: dict, snapshot_arg: str | None) -> Path:
    output_root = Path(cfg["paths"]["output_root"])
    snap_dir = output_root / "stage_c" / "snapshots"
    if not snap_dir.exists():
        raise FileNotFoundError(f"Stage C output not found at {snap_dir}.")
    parquets = sorted(snap_dir.glob("*.parquet"), key=lambda p: p.stem)
    if not parquets:
        raise FileNotFoundError(f"No snapshot parquet in {snap_dir}.")
    # Pick by --snapshot arg, else schedule.analysis_snapshot, else latest by date
    from critical_path import select_single_snapshot
    return select_single_snapshot(cfg, snapshot_arg, parquets)


# ---------------------------------------------------------------------------
# Stage classification  (priority-ordered, config-driven)
# ---------------------------------------------------------------------------

def classify_stage(activity_name: str, stage_defs: list, fallback: str) -> str:
    """First stage whose keyword (case-insensitive substring) is in the name.

    Each entry in stage_defs must be a dict like {"stage": "...", "keywords": [...]}.
    Defensive against malformed/legacy config (e.g. a bare list of stage-name
    strings): non-dict entries are skipped rather than raising, so a bad config
    degrades to the fallback stage instead of crashing the whole run."""
    nm = (activity_name or "").lower()
    for sd in stage_defs:
        if not isinstance(sd, dict):
            continue
        for kw in sd.get("keywords", []):
            if kw and kw.lower() in nm:
                return sd.get("stage", fallback)
    return fallback


# ---------------------------------------------------------------------------
# Business-day signed difference (start slip)
# ---------------------------------------------------------------------------

def business_day_slip(baseline_start, actual_start, weekmask="1111100", holidays=None):
    """Signed Mon-Fri business days from baseline start to actual start."""
    if baseline_start is None or actual_start is None \
            or pd.isna(baseline_start) or pd.isna(actual_start):
        return None
    b = np.datetime64(pd.Timestamp(baseline_start).date(), "D")
    a = np.datetime64(pd.Timestamp(actual_start).date(), "D")
    hol = np.array(holidays or [], dtype="datetime64[D]")
    if a >= b:
        return float(np.busday_count(b, a, weekmask=weekmask, holidays=hol))
    return -float(np.busday_count(a, b, weekmask=weekmask, holidays=hol))


# ---------------------------------------------------------------------------
# Sub-category list formatting
# ---------------------------------------------------------------------------

def format_subcats(subcats: list, package_name: str, show: int = 5) -> str:
    """Distinct sub-categories (excluding the package name), first `show` + (+N more)."""
    distinct = [s for s in sorted(set(subcats)) if s and s != package_name]
    if not distinct:
        return ""
    if len(distinct) <= show:
        return ", ".join(distinct)
    return ", ".join(distinct[:show]) + f"  (+{len(distinct) - show} more)"


# ---------------------------------------------------------------------------
# Core analysis  (pure, testable)
# ---------------------------------------------------------------------------

def build_buyout_analysis(grouping_df, packages_df, tasks_df, cfg):
    """Returns dict of result DataFrames + report."""
    ba = cfg["buyout_analysis"]
    stage_defs = ba["stage_classification"]
    fallback = ba.get("stage_fallback", "Other")
    top_n = ba.get("top_packages_count", 25)
    min_base = cfg["construction_variance"].get("baseline_span_min_days", 0.5)
    cal = cfg.get("working_calendar", {})
    weekmask = cal.get("weekmask", "1111100")
    holidays = cal.get("holidays", [])

    # Use the buyout-specific baseline columns (which MS Project baseline slot
    # they came from is set by schedule.buyout_baseline_number at extraction
    # time). Alias them onto the generic baseline_* names everything below reads,
    # falling back to the generic columns for parquet files from an older build.
    tasks_df = tasks_df.copy()
    for _f in ("start", "finish", "duration"):
        _src = f"buyout_baseline_{_f}"
        if _src in tasks_df.columns:
            tasks_df[f"baseline_{_f}"] = tasks_df[_src]

    # ── leaf-level join: grouping + snapshot fields ───────────────────────
    snap_cols = ["uid", "baseline_duration", "duration", "baseline_start",
                 "actual_start", "baseline_finish", "actual_finish", "pct_complete"]
    snap = tasks_df[[c for c in snap_cols if c in tasks_df.columns]].copy()
    leaf = grouping_df.merge(snap, on="uid", how="left")

    # stage + per-leaf duration variance + start slip
    leaf["stage"] = leaf["activity"].map(lambda n: classify_stage(n, stage_defs, fallback))
    leaf["base_dur"] = leaf["baseline_duration"].fillna(0.0)
    leaf["act_dur"] = leaf["duration"].fillna(0.0)
    leaf["dur_var"] = leaf["act_dur"] - leaf["base_dur"]
    leaf["start_slip"] = [
        business_day_slip(bs, as_, weekmask, holidays)
        for bs, as_ in zip(leaf["baseline_start"], leaf["actual_start"])
    ]

    # ── package-level metrics (SUMMARY-SPAN basis) ────────────────────────
    task_by_uid = tasks_df.set_index("uid")
    pkg_rows = []
    for _, pk in packages_df.iterrows():
        puid = int(pk["package_uid"])
        # package baseline/actual = the package SUMMARY TASK's own duration
        if puid in task_by_uid.index:
            prow = task_by_uid.loc[puid]
            base = float(prow["baseline_duration"]) if pd.notna(prow["baseline_duration"]) else 0.0
            act = float(prow["duration"]) if pd.notna(prow["duration"]) else 0.0
        else:
            base = act = 0.0
        var = act - base
        pct = (var / base) if base >= min_base else np.nan

        pleaves = leaf[leaf["package_uid"] == puid]
        subcats = format_subcats(list(pleaves["sub_category"]), pk["package_name"])
        astart = pleaves["actual_start"].min() if not pleaves.empty else None
        afinish = pleaves["actual_finish"].max() if not pleaves.empty else None

        pkg_rows.append({
            "package_uid":  puid,
            "section":      pk["section"],
            "group":        pk["group"],
            "phase":        pk.get("phase", "") or "",
            "category":     pk["package_name"],
            "sub_categories": subcats,
            "activities":   int(pk["leaf_count"]),
            "baseline":     round(base, 2),
            "actual":       round(act, 2),
            "abs_var":      round(var, 2),
            "pct_var":      pct,
            "actual_start": astart,
            "actual_finish": afinish,
        })
    # Explicit columns: pkg_rows is empty whenever the project has no buyout
    # scope at all (Stage D resolves zero packages). pd.DataFrame([]) would
    # drop every column, and pkg["section"]/pkg["group"]/sort_values("abs_var")
    # below are not guarded by an .empty check, so they'd raise KeyError.
    pkg_cols = ["package_uid", "section", "group", "phase", "category",
                "sub_categories", "activities", "baseline", "actual", "abs_var",
                "pct_var", "actual_start", "actual_finish"]
    pkg = pd.DataFrame(pkg_rows, columns=pkg_cols)

    # ── bucket summary (Section × Group): SUM of package spans ────────────
    bucket_rows = []
    bucket_order = [("Buyout Work", "Procurement"), ("Buyout Work", "Subcontracting"),
                    ("Lead Time", "Procurement"), ("Lead Time", "Subcontracting")]
    for sec, grp in bucket_order:
        sub = pkg[(pkg["section"] == sec) & (pkg["group"] == grp)]
        if sub.empty:
            continue
        base_sum = sub["baseline"].sum()
        act_sum = sub["actual"].sum()
        var_sum = act_sum - base_sum
        leaf_count = int(sub["activities"].sum())
        # largest package overrun in the bucket
        top_pkg = sub.loc[sub["abs_var"].idxmax()]
        ph = top_pkg["phase"]
        label = (f"{ph} ▸ {top_pkg['category']} (+{int(round(top_pkg['abs_var']))}d)"
                 if ph else f"{top_pkg['category']} (+{int(round(top_pkg['abs_var']))}d)")
        bucket_rows.append({
            "bucket":        f"{sec} – {grp}",
            "packages":      len(sub),
            "activities":    leaf_count,
            "baseline_span": round(base_sum, 2),
            "actual_span":   round(act_sum, 2),
            "abs_var":       round(var_sum, 2),
            "pct_var":       (var_sum / base_sum) if base_sum >= min_base else np.nan,
            "largest_overrun": label,
        })
    summary = pd.DataFrame(bucket_rows)
    # TOTAL row
    if not summary.empty:
        tb = summary["baseline_span"].sum()
        ta = summary["actual_span"].sum()
        total = {
            "bucket": "TOTAL — Buyout",
            "packages": int(summary["packages"].sum()),
            "activities": int(summary["activities"].sum()),
            "baseline_span": round(tb, 2),
            "actual_span": round(ta, 2),
            "abs_var": round(ta - tb, 2),
            "pct_var": ((ta - tb) / tb) if tb >= min_base else np.nan,
            "largest_overrun": "",
        }
        summary = pd.concat([summary, pd.DataFrame([total])], ignore_index=True)

    # ── top packages (ranked by abs duration variance) ────────────────────
    top_pkgs = pkg.sort_values("abs_var", ascending=False).head(top_n).reset_index(drop=True)
    top_pkgs.insert(0, "rank", range(1, len(top_pkgs) + 1))

    # ── per-bucket package tabs ───────────────────────────────────────────
    bucket_tabs = {}
    for sec, grp in bucket_order:
        sub = pkg[(pkg["section"] == sec) & (pkg["group"] == grp)] \
            .sort_values("abs_var", ascending=False).reset_index(drop=True)
        if sub.empty:
            continue
        sub.insert(0, "rank", range(1, len(sub) + 1))
        bucket_tabs[f"{sec} – {grp}"] = sub

    # ── stage breakdown (all leaves) ──────────────────────────────────────
    stage_rows = []
    stage_order = [sd["stage"] for sd in stage_defs] + [fallback]
    for st in stage_order:
        sl = leaf[leaf["stage"] == st]
        if sl.empty:
            continue
        occ = len(sl)
        avg_base = sl["base_dur"].mean()
        avg_act = sl["act_dur"].mean()
        total_var = sl["dur_var"].sum()
        stage_rows.append({
            "stage":      st,
            "occurrences": occ,
            "avg_baseline": round(avg_base, 2),
            "avg_actual":  round(avg_act, 2),
            "avg_var":     round(avg_act - avg_base, 2),
            "total_var":   round(total_var, 1),
            "median_start_slip": round(sl["start_slip"].dropna().median(), 1)
                                 if sl["start_slip"].notna().any() else None,
        })
    stage_breakdown = pd.DataFrame(stage_rows)

    # ── activity detail (every leaf) ──────────────────────────────────────
    detail = leaf[["section", "group", "phase", "category", "sub_category",
                   "activity", "stage", "base_dur", "act_dur", "dur_var",
                   "pct_complete", "baseline_finish", "actual_finish",
                   "start_slip", "is_lead_time"]].copy()
    detail = detail.rename(columns={"base_dur": "baseline", "act_dur": "actual",
                                    "dur_var": "abs_var"})

    report = {
        "buyout_leaf_count": len(leaf),
        "package_count": len(pkg),
        "stage_distribution": leaf["stage"].value_counts().to_dict(),
        "bucket_count": len(bucket_rows),
        "total_baseline_span": float(summary.iloc[-1]["baseline_span"]) if not summary.empty else 0,
        "total_actual_span": float(summary.iloc[-1]["actual_span"]) if not summary.empty else 0,
        "total_abs_var": float(summary.iloc[-1]["abs_var"]) if not summary.empty else 0,
    }

    return {
        "summary": summary, "packages": pkg, "top_packages": top_pkgs,
        "bucket_tabs": bucket_tabs, "stage_breakdown": stage_breakdown,
        "detail": detail, "report": report,
    }


# ---------------------------------------------------------------------------
# Styled workbook  (J New Town template)
# ---------------------------------------------------------------------------

NAVY = "FF1F4E78"; ZEBRA = "FFF2F6FB"; REDTINT = "FFFCE4E4"; GRAY = "FF606060"; WHITE = "FFFFFFFF"


def _d(v):
    try:
        ts = pd.Timestamp(v)
        return ts.strftime("%Y-%m-%d") if pd.notna(ts) else ""
    except Exception:
        return ""


def write_workbook(path, cfg, res):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    navy = PatternFill("solid", fgColor=NAVY); zebra = PatternFill("solid", fgColor=ZEBRA)
    red = PatternFill("solid", fgColor=REDTINT)
    hf = Font(name="Calibri", bold=True, color=WHITE)
    tf = Font(name="Calibri", size=14, bold=True, color=NAVY)
    sf = Font(name="Calibri", size=10, color=GRAY)
    nf = Font(name="Calibri"); nb = Font(name="Calibri", bold=True)
    status = cfg["project"].get("analysis_status_date", "")
    pname = cfg["project"].get("name", "")
    subtitle = ("Methodology — duration-based. Variance = Actual − Baseline working days. "
                f"Package rows use the MS Project summary-task span (parallel-safe). Status {status}.")

    def autosize(ws, headers, hr, maxw=46):
        for c in range(1, len(headers) + 1):
            L = get_column_letter(c); mx = len(str(headers[c - 1]))
            for row in ws.iter_rows(min_col=c, max_col=c, min_row=hr + 1):
                for cell in row:
                    if cell.value is not None and not str(cell.value).startswith("="):
                        mx = max(mx, min(len(str(cell.value)), 50))
            ws.column_dimensions[L].width = min(max(mx + 2, 8), maxw)

    def write_table(ws, title, headers, rows, hr=4, num_cols=None, pct_cols=None,
                    var_cols=None, formula_var=None):
        ws["A1"] = title; ws["A1"].font = tf
        ws["A2"] = subtitle; ws["A2"].font = sf
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=hr, column=c, value=h)
            cell.font = hf; cell.fill = navy
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        num_cols = num_cols or set(); pct_cols = pct_cols or set(); var_cols = var_cols or set()
        for i, rowvals in enumerate(rows):
            rr = hr + 1 + i; fill = zebra if i % 2 else None
            for c, v in enumerate(rowvals, 1):
                cell = ws.cell(row=rr, column=c, value=v)
                if c in var_cols:
                    cell.font = nb; cell.fill = red
                    cell.number_format = "0.0"; cell.alignment = Alignment(horizontal="right")
                    continue
                cell.font = nf
                if c in pct_cols:
                    cell.number_format = "0.0%"; cell.alignment = Alignment(horizontal="right")
                elif c in num_cols:
                    cell.number_format = "0.0"; cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
                if fill:
                    cell.fill = fill
        autosize(ws, headers, hr)

    def pctval(p):
        return (float(p) if pd.notna(p) else "n/a")

    wb = Workbook()

    # ── Summary ───────────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Summary"
    s = res["summary"]
    hdrs = ["Bucket", "Packages", "Activities", "Baseline Span Σ (d)",
            "Actual Span Σ (d)", "Abs. Var (d)", "% Var", "Largest Package Overrun"]
    rows = [[r["bucket"], r["packages"], r["activities"], r["baseline_span"],
             r["actual_span"], r["abs_var"], pctval(r["pct_var"]), r["largest_overrun"]]
            for _, r in s.iterrows()]
    write_table(ws, f"{pname} — Buyout Duration Variance", hdrs, rows,
                num_cols={4, 5}, pct_cols={7}, var_cols={6})
    # bold the TOTAL row
    if rows:
        for c in range(1, 9):
            ws.cell(row=4 + len(rows), column=c).font = Font(name="Calibri", bold=True)
    # how-to-read notes
    base = 4 + len(rows) + 2
    notes = [
        "How to read this workbook",
        "• Package rows (Summary, the four bucket tabs, Top Packages) use each package's MS Project summary-task duration — its scheduled working span. Headline, parallel-safe: does NOT multiply by the 16 buildings.",
        "• Stage Breakdown uses a different lens: it sums each activity occurrence's own duration overrun to show which kind of step stretched. Its totals are larger because they include per-building repetition — read Avg Var (d/occ).",
        "• Absolute working days is the primary metric. % Var is shown for J New Town consistency but is unreliable where baselines are near-zero (Lead Time, TBD); those cells show 'n/a'.",
        "• Buyout baselines are front-loaded; several scopes are unresolved 'TBD' items, so magnitudes are relative diagnostics. Lead Time entries are near-zero-duration markers whose span growth reflects date rebaselining.",
    ]
    for i, n in enumerate(notes):
        cell = ws.cell(row=base + i, column=1, value=n)
        cell.font = Font(name="Calibri", bold=(i == 0), color=NAVY if i == 0 else "FF000000")

    # ── Stage Breakdown ───────────────────────────────────────────────────
    ws = wb.create_sheet("Stage Breakdown")
    sb = res["stage_breakdown"]
    hdrs = ["Activity Stage", "Occurrences", "Avg Baseline (d)", "Avg Actual (d)",
            "Avg Var (d/occ)", "Total Var (d)*"]
    rows = [[r["stage"], r["occurrences"], r["avg_baseline"], r["avg_actual"],
             r["avg_var"], r["total_var"]] for _, r in sb.iterrows()]
    write_table(ws, f"{pname} — Buyout by Activity Stage", hdrs, rows, num_cols={3, 4, 5, 6})
    base = 4 + len(rows) + 2
    ws.cell(row=base, column=1,
            value="* Total Var sums every activity occurrence across all buildings; "
                  "use Avg Var (d/occ) for the cleanest cross-stage comparison.").font = nf

    # ── Top Packages ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Top Packages")
    tp = res["top_packages"]
    hdrs = ["Rank", "Section", "Group", "Phase", "Category (Package)", "Sub-Categories",
            "Activities", "Baseline (d)", "Actual (d)", "Abs Var (d)", "% Var"]
    rows = [[r["rank"], r["section"], r["group"], r["phase"], r["category"],
             r["sub_categories"], r["activities"], r["baseline"], r["actual"],
             r["abs_var"], pctval(r["pct_var"])] for _, r in tp.iterrows()]
    write_table(ws, f"Top {len(tp)} Buyout Packages by Duration Variance", hdrs, rows,
                num_cols={8, 9}, pct_cols={11}, var_cols={10})

    # ── Four bucket tabs ──────────────────────────────────────────────────
    tab_names = {"Buyout Work – Procurement": "Buyout Work – Proc",
                 "Buyout Work – Subcontracting": "Buyout Work – Sub",
                 "Lead Time – Procurement": "Lead Time – Proc",
                 "Lead Time – Subcontracting": "Lead Time – Sub"}
    for bucket, sheet in tab_names.items():
        if bucket not in res["bucket_tabs"]:
            continue
        bt = res["bucket_tabs"][bucket]
        ws = wb.create_sheet(sheet)
        hdrs = ["Rank", "Category (Package)", "Phase", "Sub-Categories", "Activities",
                "Baseline (d)", "Actual (d)", "Abs Var (d)", "% Var", "Actual Start", "Actual Finish"]
        rows = [[r["rank"], r["category"], r["phase"], r["sub_categories"], r["activities"],
                 r["baseline"], r["actual"], r["abs_var"], pctval(r["pct_var"]),
                 _d(r["actual_start"]), _d(r["actual_finish"])] for _, r in bt.iterrows()]
        write_table(ws, f"{bucket} — Duration Variance", hdrs, rows,
                    num_cols={6, 7}, pct_cols={9}, var_cols={8})

    # ── Activity Detail ───────────────────────────────────────────────────
    ws = wb.create_sheet("Activity Detail")
    det = res["detail"]
    hdrs = ["Section", "Group", "Phase", "Category", "Sub-Category", "Activity", "Stage",
            "Baseline (d)", "Actual (d)", "Abs Var (d)", "% Complete", "Baseline Finish", "Actual Finish"]
    rows = [[r["section"], r["group"], r["phase"], r["category"], r["sub_category"],
             r["activity"], r["stage"], round(float(r["baseline"]), 2), round(float(r["actual"]), 2),
             round(float(r["abs_var"]), 2),
             (float(r["pct_complete"]) / 100.0 if pd.notna(r["pct_complete"]) else 0.0),
             _d(r["baseline_finish"]), _d(r["actual_finish"])] for _, r in det.iterrows()]
    write_table(ws, f"{pname} — Buyout Activity Detail (every leaf)", hdrs, rows,
                num_cols={8, 9}, pct_cols={11}, var_cols={10})

    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Stage H — buyout phase analysis")
    parser.add_argument("--config", default="project_config.json")
    parser.add_argument("--snapshot", default=None)
    args = parser.parse_args()

    cfg = load_config(args.config)
    project_name = cfg["project"]["name"]
    output_root = Path(cfg["paths"]["output_root"])
    stage_dir = output_root / "stage_h"
    stage_dir.mkdir(parents=True, exist_ok=True)

    stage_d_dir = output_root / "stage_d"
    grouping_path = stage_d_dir / "buyout_grouping.parquet"
    packages_path = stage_d_dir / "buyout_packages.parquet"
    if not grouping_path.exists() or not packages_path.exists():
        sys.exit(f"Stage D output not found in {stage_d_dir}. Run resolve_wbs.py first.")

    snapshot_path = find_snapshot(cfg, args.snapshot)

    print(f"\n{'='*60}")
    print(f"  Stage H — Buyout Phase Analysis")
    print(f"  Project  : {project_name}")
    print(f"  Snapshot : {snapshot_path.stem}")
    print(f"{'='*60}\n")

    grouping_df = pd.read_parquet(grouping_path)
    packages_df = pd.read_parquet(packages_path)
    tasks_df = pd.read_parquet(snapshot_path)

    res = build_buyout_analysis(grouping_df, packages_df, tasks_df, cfg)

    # ── persist ───────────────────────────────────────────────────────────
    res["summary"].to_parquet(stage_dir / "buyout_summary.parquet", index=False)
    res["packages"].to_parquet(stage_dir / "buyout_packages_ranked.parquet", index=False)
    res["stage_breakdown"].to_parquet(stage_dir / "buyout_stage_breakdown.parquet", index=False)
    res["detail"].to_parquet(stage_dir / "buyout_activity_detail.parquet", index=False)

    xlsx_path = stage_dir / f"Buyout_Duration_Variance_{snapshot_path.stem}.xlsx"
    write_workbook(xlsx_path, cfg, res)

    report = res["report"]
    report["generated"] = datetime.now().isoformat(timespec="seconds")
    report["project"] = project_name
    report["snapshot"] = snapshot_path.stem
    with open(stage_dir / "buyout_report.json", "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)

    # ── console summary ───────────────────────────────────────────────────
    print(f"  Buyout leaves : {report['buyout_leaf_count']}")
    print(f"  Packages      : {report['package_count']}")
    print(f"\n  Stage distribution:")
    for st, c in report["stage_distribution"].items():
        print(f"    {st:<30} {c:>5}")
    print(f"\n  Bucket summary (Section – Group, summary-span Σ):")
    for _, r in res["summary"].iterrows():
        print(f"    {r['bucket']:<28} {r['packages']:>3} pkg  "
              f"base {r['baseline_span']:>8.1f}  act {r['actual_span']:>9.1f}  "
              f"var {r['abs_var']:>+9.1f}")

    print(f"\n{'='*60}")
    print(f"  Workbook : {xlsx_path}")
    print(f"  Parquet  : {stage_dir}")
    print(f"{'='*60}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
