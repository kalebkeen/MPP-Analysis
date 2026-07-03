"""
Stage K — Charting  (the seven core charts + editable chart data workbook)
============================================================================
Generates the brief's seven core charts from the Stage E/F/G parquet outputs.
Each chart takes the project name from config, applies the navy / Carlito house
style, and annotates directly on the figure where that replaces caption text.

Charts (output_root/stage_k/):
  1. forecast_trend.png        — forecast completion week-by-week vs baseline
  2. bucket_trajectories.png   — bucket variance build-up (Site Work + top buildings)
  3. driving_resource.png      — driving-path forecast, colored by controlling resource
  4. delay_waterfall.png       — chronological control-period waterfall (red add / green recover)
  5. resource_net_bar.png      — net days by controlling resource, ranked
  6. building_lollipop.png     — building turnover baseline→forecast lollipop
  7. float_histogram.png       — float-health bands

Also writes chart_data_workbook.xlsx: one tab per chart, the same underlying
data table plus a native, editable Excel chart object built from it - so an
exec can tweak numbers/formulas without touching Python. Each chart's data
assembly happens exactly once (in a _data_* function below); both the PNG
(matplotlib) and the workbook (openpyxl) render from that single dataframe,
so the two outputs can't drift apart if a metric or formula changes later.

Inputs (reads whatever is present; skips a chart with a note if its stage
output is missing):
  - Stage F: output_root/stage_f/{snapshot_controlling,waterfall_periods,by_resource_net}.parquet
  - Stage G: output_root/stage_g/{building_turnover,float_health}.parquet
  - Stage E: reused live for the per-snapshot bucket trajectory
  - Stage C: snapshots, for the trajectory

Usage:
  python generate_charts.py [--config project_config.json]
"""

import argparse
import json
import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.font_manager as fm
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties


# ---------------------------------------------------------------------------
# House style
# ---------------------------------------------------------------------------

# Optional snapshot-order manifest (set from --manifest in main), consumed by
# the multi-snapshot data functions via discover_snapshots.
MANIFEST_PATH = None

NAVY = "#1F4E78"
RED = "#C0392B"      # adds days
GREEN = "#27632A"    # recovers days
GRID = "#D9D9D9"
GRAYTXT = "#606060"
# muted supporting palette for multi-series / categorical
PALETTE = ["#1F4E78", "#E67E22", "#27632A", "#7D5BA6", "#C0392B", "#2C7BB6",
           "#9B59B6", "#16A085", "#D4AC0D", "#A0522D", "#5D6D7E", "#34495E"]

# Installer bundles the 4 Carlito TTFs into a "fonts" folder alongside this
# script (see PA-Pipeline-Setup.cs's CARLITO_FONTS_RESOURCE_NAME) - resolve
# relative to this file's own location, not a Linux-only system font path.
CARLITO_DIR = Path(__file__).resolve().parent / "fonts"


def register_fonts():
    """Register Carlito (regular + bold) and set it as the default family."""
    registered = False
    for fn in ["Carlito-Regular.ttf", "Carlito-Bold.ttf", "Carlito-Italic.ttf",
               "Carlito-BoldItalic.ttf"]:
        p = CARLITO_DIR / fn
        if p.exists():
            fm.fontManager.addfont(str(p))
            registered = True
    if not registered:
        print("  NOTE: Carlito not found; using Arial/DejaVu Sans fallback.")
    # Use a sans-serif fallback CHAIN rather than a single family name. Setting
    # font.family to a specific missing font (e.g. "Liberation Sans", which is
    # not present on stock Windows) makes matplotlib emit a "findfont: Font
    # family not found" warning for EVERY text element drawn - hundreds per run.
    # A chain ending in DejaVu Sans (always bundled with matplotlib) resolves
    # silently: Carlito if registered, else Arial on Windows, else DejaVu Sans.
    #
    # Calibri is deliberately NOT in this chain. Verified bug, reproduced in
    # isolation: at the non-default DPI this module renders at (figure.dpi /
    # savefig.dpi = 150 below), matplotlib's Agg/FreeType text path silently
    # corrupts ONLY ax.legend() text when the resolved font is Calibri - every
    # legend entry collapses to literally "ti" (title/axis-label/tick text on
    # the same figure render correctly; the in-memory Text objects still hold
    # the full correct string even after the corrupted save). Reproduced with
    # Calibri at dpi 120/150, not at the matplotlib default of 100. Confirmed
    # NOT a labeling bug in this file: passing an explicit FontProperties to
    # legend(), or forcing an extra fig.canvas.draw() before savefig(), does
    # not fix it. Arial and DejaVu Sans do not exhibit the bug at the same
    # dpi, so Arial replaces Calibri as the practical Windows fallback.
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["font.sans-serif"] = ["Carlito", "Arial", "Liberation Sans", "DejaVu Sans"]
    plt.rcParams.update({
        "font.size": 10,
        "axes.titlesize": 13,
        "axes.titleweight": "bold",
        "axes.titlecolor": NAVY,
        "axes.labelcolor": GRAYTXT,
        "axes.edgecolor": "#BFBFBF",
        "axes.grid": True,
        "grid.color": GRID,
        "grid.linewidth": 0.6,
        "xtick.color": GRAYTXT,
        "ytick.color": GRAYTXT,
        "figure.dpi": 150,
        "savefig.dpi": 150,
        "savefig.bbox": "tight",
    })


def load_config(config_path: str) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _to_dt(series):
    return pd.to_datetime(series, errors="coerce")


def _save(fig, path):
    fig.savefig(path, facecolor="white")
    plt.close(fig)
    print(f"    saved {path.name}")


def _xlsx_font_name(cfg):
    return cfg.get("charting", {}).get("font_family") or "Calibri"


# ---------------------------------------------------------------------------
# 1. Forecast completion trend vs baseline
# ---------------------------------------------------------------------------

def _data_forecast_trend(cfg, paths, project):
    src = paths["stage_f"] / "snapshot_controlling.parquet"
    if not src.exists():
        print("    [skip] forecast_trend — Stage F output missing")
        return None, None
    df = pd.read_parquet(src)
    df = df[df["milestone_finish"].notna()].copy()
    df["Snapshot Date"] = _to_dt(df["date"])
    df["Forecast Completion"] = _to_dt(df["milestone_finish"])
    df = df.sort_values("Snapshot Date")[["Snapshot Date", "Forecast Completion"]].reset_index(drop=True)

    baseline = cfg["schedule"].get("baseline_date")
    df["Baseline Completion"] = pd.Timestamp(baseline) if baseline else pd.NaT

    title = f"{project} — Forecast Completion Date, Week by Week"
    return df, title


def _render_forecast_trend(df, title, cfg, project, out_dir):
    fig, ax = plt.subplots(figsize=(10, 5.2))
    ax.plot(df["Snapshot Date"], df["Forecast Completion"], color=NAVY, lw=2, marker="o", ms=4,
            label="Forecast completion (each weekly snapshot)")

    if df["Baseline Completion"].notna().any():
        bdt = df["Baseline Completion"].iloc[0]
        ax.axhline(bdt, color=RED, ls="--", lw=1.5,
                   label=f"Baseline completion ({bdt.strftime('%b %#d, %Y')})")

    for bp in cfg.get("charting", {}).get("forecast_trend_chart", {}).get("annotation_breakpoints", []):
        try:
            bpd = pd.Timestamp(bp["date"])
            row = df.iloc[(df["Snapshot Date"] - bpd).abs().argsort().iloc[0]]
            ax.annotate(bp.get("label", ""), xy=(row["Snapshot Date"], row["Forecast Completion"]),
                        xytext=(0, 28), textcoords="offset points", ha="center",
                        fontsize=8, color=GRAYTXT,
                        arrowprops=dict(arrowstyle="->", color=GRAYTXT, lw=0.8))
        except Exception:
            continue

    ax.set_title(title)
    ax.set_ylabel("Forecast completion date")
    ax.set_xlabel("Snapshot date")
    ax.yaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.legend(loc="upper left", fontsize=8, framealpha=0.9)
    fig.autofmt_xdate(rotation=0, ha="center")
    _save(fig, out_dir / "forecast_trend.png")


# ---------------------------------------------------------------------------
# 2. Bucket variance trajectories  (reuses Stage E per snapshot)
# ---------------------------------------------------------------------------

def _data_bucket_trajectories(cfg, paths, project):
    try:
        from construction_variance import build_construction_variance
        from critical_path import discover_snapshots
    except ImportError:
        print("    [skip] bucket_trajectories — needs construction_variance.py + critical_path.py")
        return None, None
    try:
        # MANIFEST_PATH mirrors Stage F/G's --manifest: without it, a snapshot
        # whose filename carries no parseable date silently orders by file
        # mtime, which is the COM-export date, not the snapshot date.
        snaps = discover_snapshots(cfg, MANIFEST_PATH)
    except Exception as e:
        print(f"    [skip] bucket_trajectories — {e}")
        return None, None

    # compute each snapshot's per-bucket net variance
    records = []
    for snap in snaps:
        try:
            tdf = pd.read_parquet(snap["tasks_path"])
            _, _, _, bsum, _ = build_construction_variance(tdf, cfg)
            if bsum.empty:
                continue
            for _, r in bsum.iterrows():
                records.append({"date": snap["date"], "bucket": r["bucket"],
                                "net": r["net_variance"]})
        except Exception:
            continue
    if not records:
        print("    [skip] bucket_trajectories — no bucket data computed")
        return None, None
    traj = pd.DataFrame(records)
    traj["date"] = _to_dt(traj["date"])

    # pick Site Work + the top building buckets by latest absolute variance
    latest = traj.sort_values("date").groupby("bucket").last()
    building_names = set(cfg["buildings"]["names"])
    site = [b for b in latest.index if b not in building_names]
    top_buildings = (latest[latest.index.isin(building_names)]["net"].abs()
                     .sort_values(ascending=False).head(4).index.tolist())
    keep = [b for b in ["Site Work"] if b in latest.index] + \
           [b for b in site if b != "Site Work"][:2] + top_buildings
    keep = list(dict.fromkeys(keep))

    wide = traj[traj["bucket"].isin(keep)].pivot_table(index="date", columns="bucket", values="net")
    wide = wide.reindex(columns=[b for b in keep if b in wide.columns]).sort_index().reset_index()
    wide = wide.rename(columns={"date": "Snapshot Date"})

    title = f"{project} — Variance Build-Up by Bucket"
    return wide, title


def _render_bucket_trajectories(df, title, cfg, project, out_dir):
    series_cols = [c for c in df.columns if c != "Snapshot Date"]
    fig, ax = plt.subplots(figsize=(10, 5.2))
    for i, b in enumerate(series_cols):
        sub = df[["Snapshot Date", b]].dropna()
        ax.plot(sub["Snapshot Date"], sub[b], lw=1.8, marker="o", ms=3,
                color=PALETTE[i % len(PALETTE)], label=b)
    ax.axhline(0, color="#999999", lw=0.8)
    ax.set_title(title)
    ax.set_ylabel("Net bucket variance (working days)")
    ax.set_xlabel("Snapshot date")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.legend(loc="upper left", fontsize=8, ncol=2, framealpha=0.9)
    fig.autofmt_xdate(rotation=0, ha="center")
    _save(fig, out_dir / "bucket_trajectories.png")


# ---------------------------------------------------------------------------
# 3. Driving-resource colored scatter
# ---------------------------------------------------------------------------

def _data_driving_resource(cfg, paths, project):
    src = paths["stage_f"] / "snapshot_controlling.parquet"
    if not src.exists():
        print("    [skip] driving_resource — Stage F output missing")
        return None, None
    df = pd.read_parquet(src)
    df = df[df["milestone_finish"].notna()].copy()
    df["Snapshot Date"] = _to_dt(df["date"])
    df["Forecast Completion"] = _to_dt(df["milestone_finish"])
    df["res"] = df["controlling_resources"].fillna("(unassigned)").replace("", "(unassigned)")
    df = df.sort_values("Snapshot Date")

    # resources controlling more than one week get a color; singletons -> "other".
    # Legend cap (5.10): above legend_max_series distinct resources, only the
    # top N by weeks-in-control keep their own series — the rest group into
    # "other", so a 30-sub project can't explode the legend off the chart.
    counts = df["res"].value_counts()
    main = counts[counts > 1].index.tolist()
    max_series = int(cfg.get("charting", {}).get("legend_max_series", 10) or 10)
    if len(main) > max_series:
        main = counts.index.tolist()[:max_series]
    df["res_group"] = df["res"].where(df["res"].isin(main), "other (1 wk each)")

    pivot = df.pivot_table(index="Snapshot Date", columns="res_group",
                            values="Forecast Completion", aggfunc="first")
    order = main + (["other (1 wk each)"] if "other (1 wk each)" in pivot.columns else [])
    pivot = pivot.reindex(columns=[c for c in order if c in pivot.columns])

    base = (df[["Snapshot Date", "Forecast Completion"]]
            .drop_duplicates("Snapshot Date").set_index("Snapshot Date"))
    wide = base.join(pivot).sort_index().reset_index()
    wide = wide.rename(columns={"Forecast Completion": "Forecast Completion (all)"})

    baseline = cfg["schedule"].get("baseline_date")
    wide["Baseline Completion"] = pd.Timestamp(baseline) if baseline else pd.NaT

    title = f"{project} — Which Resource Controls the Finish Date, Week by Week"
    return wide, title


def _render_driving_resource(df, title, cfg, project, out_dir):
    resource_cols = [c for c in df.columns
                      if c not in ("Snapshot Date", "Forecast Completion (all)", "Baseline Completion")]
    fig, ax = plt.subplots(figsize=(10, 5.2))
    ax.plot(df["Snapshot Date"], df["Forecast Completion (all)"], color="#BBBBBB", lw=1, zorder=1)
    for i, r in enumerate(resource_cols):
        sub = df[["Snapshot Date", r]].dropna()
        if r == "other (1 wk each)":
            ax.scatter(sub["Snapshot Date"], sub[r], s=30, color="#AAB7B8",
                       edgecolor="white", lw=0.5, label=r, zorder=2)
        else:
            ax.scatter(sub["Snapshot Date"], sub[r], s=42, color=PALETTE[i % len(PALETTE)],
                       edgecolor="white", lw=0.5, label=r, zorder=3)

    if df["Baseline Completion"].notna().any():
        ax.axhline(df["Baseline Completion"].iloc[0], color=RED, ls="--", lw=1.2, zorder=1)

    ax.set_title(title)
    ax.set_ylabel("Forecast completion (driving path)")
    ax.set_xlabel("Snapshot date")
    ax.yaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.legend(loc="upper left", fontsize=7.5, ncol=2, title="Controlling resource",
              title_fontsize=8, framealpha=0.9)
    fig.autofmt_xdate(rotation=0, ha="center")
    _save(fig, out_dir / "driving_resource.png")


# ---------------------------------------------------------------------------
# 4. Delay-ledger waterfall
# ---------------------------------------------------------------------------

def _data_delay_waterfall(cfg, paths, project):
    src = paths["stage_f"] / "waterfall_periods.parquet"
    if not src.exists():
        print("    [skip] delay_waterfall — Stage F output missing")
        return None, None
    df = pd.read_parquet(src)
    if df.empty:
        print("    [skip] delay_waterfall — no control-periods")
        return None, None
    out = pd.DataFrame({
        "Controlling Activity": df["controlling_name"].astype(str),
        "Net Days": df["net_days"],
    }).reset_index(drop=True)
    out["Cumulative Days"] = out["Net Days"].cumsum()
    title = f"{project} — Delay Ledger: Days Added to the Finish, by Controlling Activity"
    return out, title


def _render_delay_waterfall(df, title, cfg, project, out_dir):
    fig, ax = plt.subplots(figsize=(11, 5.6))
    running = 0
    for i, delta in enumerate(df["Net Days"]):
        color = RED if delta > 0 else (GREEN if delta < 0 else "#999999")
        bottom = running if delta >= 0 else running + delta
        ax.bar(i, abs(delta), bottom=bottom, color=color, width=0.7,
               edgecolor="white", lw=0.4)
        running += delta
    ax.axhline(running, color=NAVY, ls="--", lw=1.2)
    ax.annotate(f"net {'+' if running>=0 else ''}{running} days",
                xy=(len(df) - 1, running), xytext=(0, 8),
                textcoords="offset points", ha="right", fontsize=9,
                color=NAVY, fontweight="bold")
    ax.axhline(0, color="#666666", lw=0.8)

    labels = [n[:26] for n in df["Controlling Activity"]]
    ax.set_xticks(range(len(df)))
    ax.set_xticklabels(labels, rotation=90, fontsize=6.5, ha="center")
    ax.set_title(title)
    ax.set_ylabel("Calendar days added (red) / recovered (green)")
    _save(fig, out_dir / "delay_waterfall.png")


# ---------------------------------------------------------------------------
# 5. Net days by controlling resource  (horizontal bar)
# ---------------------------------------------------------------------------

def _data_resource_net_bar(cfg, paths, project):
    src = paths["stage_f"] / "by_resource_net.parquet"
    if not src.exists():
        print("    [skip] resource_net_bar — Stage F output missing")
        return None, None
    df = pd.read_parquet(src)
    if df.empty:
        print("    [skip] resource_net_bar — no resource data")
        return None, None
    df = df.reindex(df["net_days"].abs().sort_values(ascending=True).index)
    out = pd.DataFrame({
        "Controlling Resource": df["controlling_resources"],
        "Net Days": df["net_days"],
    }).reset_index(drop=True)
    title = f"{project} — Net Finish-Date Movement by Controlling Resource"
    return out, title


def _render_resource_net_bar(df, title, cfg, project, out_dir):
    fig, ax = plt.subplots(figsize=(9, max(4, 0.4 * len(df) + 1)))
    colors = [RED if v > 0 else (GREEN if v < 0 else "#999999") for v in df["Net Days"]]
    ax.barh(df["Controlling Resource"], df["Net Days"], color=colors,
            edgecolor="white", lw=0.4)
    ax.axvline(0, color="#666666", lw=0.8)
    for y, v in enumerate(df["Net Days"]):
        ax.annotate(f"{'+' if v>0 else ''}{int(v)}",
                    xy=(v, y), xytext=(4 if v >= 0 else -4, 0),
                    textcoords="offset points", va="center",
                    ha="left" if v >= 0 else "right", fontsize=8, color=GRAYTXT)
    ax.set_title(title)
    ax.set_xlabel("Net calendar days contributed to finish (contemporaneous)")
    ax.grid(axis="y", visible=False)
    _save(fig, out_dir / "resource_net_bar.png")


# ---------------------------------------------------------------------------
# 6. Building turnover lollipop
# ---------------------------------------------------------------------------

def _data_building_lollipop(cfg, paths, project):
    src = paths["stage_g"] / "building_turnover.parquet"
    if not src.exists():
        print("    [skip] building_lollipop — Stage G output missing")
        return None, None
    df = pd.read_parquet(src)
    df = df[df["baseline_finish"].notna() & df["forecast_finish"].notna()].copy()
    if df.empty:
        print("    [skip] building_lollipop — no turnover data")
        return None, None
    out = pd.DataFrame({
        "Building": df["building"],
        "Baseline Finish": _to_dt(df["baseline_finish"]),
        "Forecast Finish": _to_dt(df["forecast_finish"]),
    })
    out["Slip Days"] = (out["Forecast Finish"] - out["Baseline Finish"]).dt.days
    out = out.sort_values("Slip Days").reset_index(drop=True)
    title = f"{project} — Building Turnover: Baseline vs Forecast"
    return out, title


def _render_building_lollipop(df, title, cfg, project, out_dir):
    fig, ax = plt.subplots(figsize=(10, max(4.5, 0.34 * len(df) + 1)))
    y = range(len(df))
    ax.hlines(y, df["Baseline Finish"], df["Forecast Finish"], color="#C9C9C9", lw=2, zorder=1)
    ax.scatter(df["Baseline Finish"], y, color="#9AA7B1", s=45, zorder=2, label="Baseline turnover")
    ax.scatter(df["Forecast Finish"], y, color=RED, s=55, zorder=3, label="Forecast turnover")
    for yi, (_, r) in zip(y, df.iterrows()):
        ax.annotate(f"+{int(r['Slip Days'])}d", xy=(r["Forecast Finish"], yi), xytext=(8, 0),
                    textcoords="offset points", va="center", fontsize=7.5, color=RED)
    ax.set_yticks(list(y))
    ax.set_yticklabels(df["Building"])

    baseline = cfg["schedule"].get("baseline_date")
    if baseline:
        ax.axvline(pd.Timestamp(baseline), color=GRAYTXT, ls=":", lw=1,
                   label=f"baseline {pd.Timestamp(baseline).strftime('%b %#d')}")

    ax.set_title(title)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.legend(loc="lower right", fontsize=8, framealpha=0.9)
    ax.grid(axis="y", visible=False)
    _save(fig, out_dir / "building_lollipop.png")


# ---------------------------------------------------------------------------
# 7. Float-health histogram
# ---------------------------------------------------------------------------

def _data_float_histogram(cfg, paths, project):
    src = paths["stage_g"] / "float_health.parquet"
    if not src.exists():
        print("    [skip] float_histogram — Stage G output missing")
        return None, None
    df = pd.read_parquet(src)
    if df.empty:
        print("    [skip] float_histogram — no float data")
        return None, None
    out = pd.DataFrame({
        "Float Band": df["band"],
        "Task Count": df["count"],
    }).reset_index(drop=True)
    title = f"{project} — Float Health: Remaining Cushion"
    return out, title


def _render_float_histogram(df, title, cfg, project, out_dir):
    colors = [RED if "Critical" in str(b) else NAVY for b in df["Float Band"]]

    fig, ax = plt.subplots(figsize=(9, 5))
    bars = ax.bar(df["Float Band"], df["Task Count"], color=colors, edgecolor="white", lw=0.5)
    total = df["Task Count"].sum()
    for bar, (_, r) in zip(bars, df.iterrows()):
        pct = r["Task Count"] / total * 100 if total else 0
        ax.annotate(f"{int(r['Task Count'])}\n{pct:.0f}%",
                    xy=(bar.get_x() + bar.get_width() / 2, bar.get_height()),
                    xytext=(0, 4), textcoords="offset points", ha="center",
                    fontsize=8, color=GRAYTXT)
    ax.set_title(title)
    ax.set_ylabel("Incomplete construction tasks")
    ax.set_xlabel("Total slack band (working days)")
    ax.grid(axis="x", visible=False)
    _save(fig, out_dir / "float_histogram.png")


# ---------------------------------------------------------------------------
# 8. S-curve — planned vs actual percent complete  (Part IV, work plan 5.8/5.10)
# ---------------------------------------------------------------------------

def _data_s_curve(cfg, paths, project):
    src = paths["stage_g"] / "progress_curve.parquet"
    if not src.exists():
        print("    [skip] s_curve — Stage G progress_curve missing")
        return None, None
    df = pd.read_parquet(src)
    if df.empty:
        print("    [skip] s_curve — no progress data")
        return None, None
    out = pd.DataFrame({
        "Snapshot Date": _to_dt(df["date"]),
        "Planned % Complete": df["planned_pct"].round(1),
        "Actual % Complete": df["actual_pct"].round(1),
    }).sort_values("Snapshot Date").reset_index(drop=True)
    title = f"{project} — S-Curve: Planned vs Actual Percent Complete"
    return out, title


def _render_s_curve(df, title, cfg, project, out_dir):
    fig, ax = plt.subplots(figsize=(10, 5.2))
    ax.plot(df["Snapshot Date"], df["Planned % Complete"], color="#888888",
            lw=1.8, ls="--", label="Planned (baseline windows)")
    ax.plot(df["Snapshot Date"], df["Actual % Complete"], color=NAVY,
            lw=2, marker="o", ms=3.5, label="Actual (each snapshot)")
    ax.set_title(title)
    ax.set_ylabel("Percent complete (baseline-duration weighted)")
    ax.set_xlabel("Snapshot date")
    ax.set_ylim(0, 105)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.legend(loc="lower right", fontsize=8, framealpha=0.9)
    fig.autofmt_xdate(rotation=0, ha="center")
    _save(fig, out_dir / "s_curve.png")


# ---------------------------------------------------------------------------
# 9. Float-erosion trend  (Part IV, work plan 5.10)
# ---------------------------------------------------------------------------

def _data_float_erosion(cfg, paths, project):
    try:
        from critical_path import discover_snapshots, is_buyout_outline
    except ImportError:
        print("    [skip] float_erosion — needs critical_path.py")
        return None, None
    try:
        snaps = discover_snapshots(cfg, MANIFEST_PATH)
    except Exception as e:
        print(f"    [skip] float_erosion — {e}")
        return None, None
    buyout_prefixes = cfg["schedule"]["buyout_outline_prefixes"]
    thresh = cfg["schedule"].get("percent_complete_threshold_complete", 100)
    rows = []
    for snap in snaps:
        try:
            df = pd.read_parquet(snap["tasks_path"],
                                 columns=["outline_number", "is_summary",
                                          "pct_complete", "total_slack"])
        except Exception:
            continue
        m = (~df["is_summary"]) & (df["pct_complete"].fillna(0) < thresh) & \
            df["total_slack"].notna() & \
            (~df["outline_number"].astype(str).map(
                lambda s: is_buyout_outline(s, buyout_prefixes)))
        sub = df[m]
        if sub.empty:
            continue
        rows.append({"Snapshot Date": pd.Timestamp(snap["date"]),
                     "Median Total Float (wd)": float(sub["total_slack"].median()),
                     "25th Percentile Float (wd)": float(sub["total_slack"].quantile(0.25))})
    if not rows:
        print("    [skip] float_erosion — no float data")
        return None, None
    out = pd.DataFrame(rows).sort_values("Snapshot Date").reset_index(drop=True)
    title = f"{project} — Float Erosion: Remaining Cushion Over Time"
    return out, title


def _render_float_erosion(df, title, cfg, project, out_dir):
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(df["Snapshot Date"], df["Median Total Float (wd)"], color=NAVY,
            lw=2, marker="o", ms=3.5, label="Median total float (incomplete tasks)")
    ax.plot(df["Snapshot Date"], df["25th Percentile Float (wd)"], color="#E67E22",
            lw=1.6, ls="-.", label="25th percentile")
    ax.axhline(0, color=RED, ls="--", lw=1.2, label="Zero float (critical)")
    ax.set_title(title)
    ax.set_ylabel("Total float (working days)")
    ax.set_xlabel("Snapshot date")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.legend(loc="best", fontsize=8, framealpha=0.9)
    fig.autofmt_xdate(rotation=0, ha="center")
    _save(fig, out_dir / "float_erosion.png")


# ---------------------------------------------------------------------------
# 10. Two-measurement scatter — start slip vs duration variance  (Part VI, 5.10)
# ---------------------------------------------------------------------------

def _data_buyout_scatter(cfg, paths, project):
    det_p = paths["stage_h"] / "buyout_activity_detail.parquet"
    pkg_p = paths["stage_h"] / "buyout_packages_ranked.parquet"
    if not det_p.exists() or not pkg_p.exists():
        print("    [skip] buyout_scatter — Stage H output missing")
        return None, None
    det = pd.read_parquet(det_p)
    pkg = pd.read_parquet(pkg_p)
    if det.empty or pkg.empty:
        print("    [skip] buyout_scatter — no buyout scope")
        return None, None
    slip = (det.dropna(subset=["start_slip"])
            .groupby("category")["start_slip"].median().rename("Start Slip (wd)"))
    out = (pkg[["category", "abs_var"]]
           .rename(columns={"category": "Package", "abs_var": "Duration Variance (wd)"})
           .merge(slip, left_on="Package", right_index=True, how="inner")
           [["Package", "Start Slip (wd)", "Duration Variance (wd)"]]
           .reset_index(drop=True))
    if out.empty:
        print("    [skip] buyout_scatter — no packages with start-slip data")
        return None, None
    title = f"{project} — Buyout Packages: Start Slip vs Duration Variance"
    return out, title


def _render_buyout_scatter(df, title, cfg, project, out_dir):
    fig, ax = plt.subplots(figsize=(9.5, 6))
    ax.scatter(df["Start Slip (wd)"], df["Duration Variance (wd)"], s=42,
               color=NAVY, edgecolor="white", lw=0.5, zorder=3)
    ax.axhline(0, color="#999999", lw=1)
    ax.axvline(0, color="#999999", lw=1)
    # quadrant annotations — the central buyout insight at a glance
    xmax = max(abs(df["Start Slip (wd)"].max()), abs(df["Start Slip (wd)"].min()), 1)
    ymax = max(abs(df["Duration Variance (wd)"].max()), abs(df["Duration Variance (wd)"].min()), 1)
    q = dict(fontsize=8, color=GRAYTXT, style="italic", ha="center")
    ax.text(xmax * 0.55,  ymax * 0.92, "started late, ran long", **q)
    ax.text(-xmax * 0.55, ymax * 0.92, "started fine, ran long", **q)
    ax.text(xmax * 0.55,  -ymax * 0.92, "started late, ran fine", **q)
    ax.text(-xmax * 0.55, -ymax * 0.92, "started fine, ran fine", **q)
    # label the biggest outliers only (readability over completeness)
    mag = (df["Start Slip (wd)"].abs() / xmax + df["Duration Variance (wd)"].abs() / ymax)
    for _, r in df.loc[mag.nlargest(6).index].iterrows():
        ax.annotate(str(r["Package"])[:24], xy=(r["Start Slip (wd)"], r["Duration Variance (wd)"]),
                    xytext=(4, 4), textcoords="offset points", fontsize=7, color=GRAYTXT)
    ax.set_title(title)
    ax.set_xlabel("Start-date slip (working days, median across package activities)")
    ax.set_ylabel("Duration variance (working days, summary-span)")
    _save(fig, out_dir / "buyout_scatter.png")


# ---------------------------------------------------------------------------
# 11. PO-cycle trend  (Part VI, 5.10 — the headline finding gets its own chart)
# ---------------------------------------------------------------------------

def _data_po_cycle(cfg, paths, project):
    det_p = paths["stage_h"] / "buyout_activity_detail.parquet"
    if not det_p.exists():
        print("    [skip] po_cycle — Stage H output missing")
        return None, None
    det = pd.read_parquet(det_p)
    if det.empty:
        print("    [skip] po_cycle — no buyout scope")
        return None, None
    po = det[det["stage"].astype(str).str.contains("order", case=False)
             & det["actual_finish"].notna()].copy()
    if len(po) < 5:
        print("    [skip] po_cycle — too few purchase-order activities")
        return None, None
    po["Finish Date"] = _to_dt(po["actual_finish"])
    po = po.sort_values("Finish Date")
    out = pd.DataFrame({
        "Finish Date": po["Finish Date"],
        "PO Duration (wd)": po["actual"].astype(float).round(1),
    }).reset_index(drop=True)
    out["Rolling Median (wd)"] = (out["PO Duration (wd)"]
                                  .rolling(window=max(5, len(out) // 10),
                                           min_periods=3, center=True).median())
    title = f"{project} — Purchase-Order Cycle Duration Over Time"
    return out, title


def _render_po_cycle(df, title, cfg, project, out_dir):
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.scatter(df["Finish Date"], df["PO Duration (wd)"], s=14, color=NAVY,
               alpha=0.45, edgecolor="none", label="PO-cycle activity", zorder=2)
    ax.plot(df["Finish Date"], df["Rolling Median (wd)"], color=RED, lw=2,
            label="Rolling median", zorder=3)
    ax.set_title(title)
    ax.set_ylabel("Actual duration (working days)")
    ax.set_xlabel("Activity finish date")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.legend(loc="best", fontsize=8, framealpha=0.9)
    fig.autofmt_xdate(rotation=0, ha="center")
    _save(fig, out_dir / "po_cycle_trend.png")


# ---------------------------------------------------------------------------
# Editable chart data workbook (stage_k/chart_data_workbook.xlsx)
# ---------------------------------------------------------------------------

# One data-assembly path per chart (the _data_* functions above) feeds both
# the matplotlib PNG (via _render_*) and this workbook - never a second,
# separately-maintained copy of a chart's data logic.
CHART_SPECS = [
    ("forecast_trend",      _data_forecast_trend,      _render_forecast_trend,      "line"),
    ("bucket_trajectories", _data_bucket_trajectories, _render_bucket_trajectories, "line"),
    ("driving_resource",    _data_driving_resource,    _render_driving_resource,    "scatter"),
    ("delay_waterfall",     _data_delay_waterfall,      _render_delay_waterfall,    "bar_pointcolor_v"),
    ("resource_net_bar",    _data_resource_net_bar,     _render_resource_net_bar,   "bar_pointcolor_h"),
    ("building_lollipop",   _data_building_lollipop,    _render_building_lollipop,  "bar_floating_h"),
    ("float_histogram",     _data_float_histogram,      _render_float_histogram,    "bar_pointcolor_v"),
    # Work plan 5.10 additions — same dual-output pattern, same house styling.
    # The two Part VI charts skip themselves cleanly when buyout is out of scope
    # (their Stage H source parquets are empty-but-valid in that state).
    ("s_curve",             _data_s_curve,              _render_s_curve,            "line"),
    ("float_erosion",       _data_float_erosion,        _render_float_erosion,      "line"),
    ("buyout_scatter",      _data_buyout_scatter,       _render_buyout_scatter,     "scatter"),
    ("po_cycle_trend",      _data_po_cycle,             _render_po_cycle,           "scatter"),
]


def _sheet_title(title):
    # Excel sheet names: 31-char max, no : \ / ? * [ ]
    t = (title or "Chart").replace(":", "-")
    for ch in "\\/?*[]":
        t = t.replace(ch, "-")
    return t[:31] if t else "Chart"


def _write_dataframe(ws, df, font_name):
    header_font = Font(name=font_name, bold=True, color="1F4E78")
    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=str(col))
        cell.font = header_font
    for r, (_, row) in enumerate(df.iterrows(), start=2):
        for c, col in enumerate(df.columns, start=1):
            val = row[col]
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime() if pd.notna(val) else None
            elif pd.isna(val):
                val = None
            ws.cell(row=r, column=c, value=val)
    return len(df), len(df.columns)


def _point_colors(values):
    out = []
    for v in values:
        if v > 0:
            out.append(RED.lstrip("#"))
        elif v < 0:
            out.append(GREEN.lstrip("#"))
        else:
            out.append("999999")
    return out


def _apply_point_colors(series, hex_colors):
    points = []
    for idx, hexcolor in enumerate(hex_colors):
        dp = DataPoint(idx=idx)
        dp.graphicalProperties = GraphicalProperties(solidFill=hexcolor)
        points.append(dp)
    series.data_points = points


def _add_chart_line(ws, n_rows, n_cols, title):
    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.y_axis.title = None
    chart.x_axis.title = None
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
    for c in range(2, n_cols + 1):
        data = Reference(ws, min_col=c, min_row=1, max_row=n_rows + 1)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for i, series in enumerate(chart.series):
        series.graphicalProperties.line.solidFill = PALETTE[i % len(PALETTE)].lstrip("#")
        series.graphicalProperties.line.width = 18000
        series.marker.symbol = "circle"
        series.marker.size = 5
    chart.height, chart.width = 11, 22
    ws.add_chart(chart, f"{get_column_letter(n_cols + 2)}2")


def _add_chart_scatter(ws, n_rows, n_cols, title):
    chart = ScatterChart()
    chart.title = title
    chart.style = 2
    chart.x_axis.title = None
    chart.y_axis.title = None
    x_ref = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
    for c in range(2, n_cols + 1):
        y_ref = Reference(ws, min_col=c, min_row=1, max_row=n_rows + 1)
        series = Series(y_ref, x_ref, title_from_data=True)
        series.marker.symbol = "circle"
        series.marker.size = 6
        series.graphicalProperties.line.noFill = True
        series.graphicalProperties.solidFill = PALETTE[(c - 2) % len(PALETTE)].lstrip("#")
        chart.series.append(series)
    chart.height, chart.width = 11, 22
    ws.add_chart(chart, f"{get_column_letter(n_cols + 2)}2")


def _add_chart_bar_pointcolor(ws, n_rows, n_cols, title, values, orientation):
    chart = BarChart()
    chart.type = "bar" if orientation == "h" else "col"
    chart.title = title
    chart.style = 10
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
    data = Reference(ws, min_col=2, min_row=1, max_row=n_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _apply_point_colors(chart.series[0], _point_colors(values))
    chart.height, chart.width = 11, 22
    ws.add_chart(chart, f"{get_column_letter(n_cols + 2)}2")


def _add_chart_bar_floating(ws, n_rows, n_cols, title, base_col, span_col):
    # "Floating bar" Gantt technique: an invisible base series (Baseline
    # Finish, a date serial) stacked under a visible span series (Slip
    # Days). Excel adds them, so the visible bar segment runs exactly from
    # the baseline date to the forecast date with no separate arithmetic
    # to keep in sync.
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = title
    chart.style = 10
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
    base = Reference(ws, min_col=base_col, min_row=1, max_row=n_rows + 1)
    span = Reference(ws, min_col=span_col, min_row=1, max_row=n_rows + 1)
    chart.add_data(base, titles_from_data=True)
    chart.add_data(span, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties = GraphicalProperties(noFill=True)
    chart.series[1].graphicalProperties = GraphicalProperties(solidFill=RED.lstrip("#"))
    chart.x_axis.number_format = "mmm yyyy"
    chart.height, chart.width = 11, 22
    ws.add_chart(chart, f"{get_column_letter(n_cols + 2)}2")


def write_chart_workbook(results, cfg, out_dir):
    font_name = _xlsx_font_name(cfg)
    wb = Workbook()
    wb.remove(wb.active)

    written = 0
    for name, df, title, kind in results:
        if df is None or df.empty:
            continue
        ws = wb.create_sheet(title=_sheet_title(title))
        n_rows, n_cols = _write_dataframe(ws, df, font_name)

        if kind == "line":
            _add_chart_line(ws, n_rows, n_cols, title)
        elif kind == "scatter":
            _add_chart_scatter(ws, n_rows, n_cols, title)
        elif kind == "bar_pointcolor_v":
            _add_chart_bar_pointcolor(ws, n_rows, n_cols, title, list(df.iloc[:, 1]), "v")
        elif kind == "bar_pointcolor_h":
            _add_chart_bar_pointcolor(ws, n_rows, n_cols, title, list(df.iloc[:, 1]), "h")
        elif kind == "bar_floating_h":
            # Building / Baseline Finish / Forecast Finish / Slip Days
            _add_chart_bar_floating(ws, n_rows, n_cols, title, base_col=2, span_col=4)
        written += 1

    if written == 0:
        print("    [skip] chart_data_workbook — no chart data available")
        return

    out_path = out_dir / "chart_data_workbook.xlsx"
    wb.save(out_path)
    print(f"    saved {out_path.name} ({written} chart{'s' if written != 1 else ''})")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    global MANIFEST_PATH
    parser = argparse.ArgumentParser(description="Stage K — generate the seven core charts")
    parser.add_argument("--config", default="project_config.json")
    parser.add_argument("--manifest", default=None,
                        help="Optional snapshot-order CSV (stem,date), same as Stage F/G")
    args = parser.parse_args()
    MANIFEST_PATH = args.manifest

    cfg = load_config(args.config)
    project = cfg["project"]["name"]
    output_root = Path(cfg["paths"]["output_root"])
    out_dir = output_root / "stage_k"
    out_dir.mkdir(parents=True, exist_ok=True)

    paths = {
        "stage_e": output_root / "stage_e",
        "stage_f": output_root / "stage_f",
        "stage_g": output_root / "stage_g",
        "stage_h": output_root / "stage_h",
        "stage_c": output_root / "stage_c",
    }

    print(f"\n{'='*60}")
    print(f"  Stage K — Charting")
    print(f"  Project : {project}")
    print(f"  Output  : {out_dir}")
    print(f"{'='*60}\n")

    register_fonts()

    print("  Generating charts:")
    results = []
    for name, data_fn, render_fn, kind in CHART_SPECS:
        df, title = data_fn(cfg, paths, project)
        if df is not None:
            render_fn(df, title, cfg, project, out_dir)
        results.append((name, df, title, kind))

    write_chart_workbook(results, cfg, out_dir)

    print(f"\n{'='*60}")
    print(f"  Charts written to {out_dir}")
    print(f"{'='*60}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
